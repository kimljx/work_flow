from __future__ import annotations

"""后端核心接口路由。

模块聚合认证、用户、任务、模板、通知、邮件、延期审批、导入导出与审计接口，
并承担接口层数据组装、权限校验与错误提示的职责。
"""

import csv
import hashlib
import io
import json
import logging
import threading
from datetime import datetime, timedelta
from pathlib import Path

import jwt
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from sqlalchemy.orm import Session

from app.config import settings
from app.constants import (
    ADMIN_ROLES,
    DELAY_STATUS_LABELS,
    MAIL_ACTION_STATUS_LABELS,
    MAIL_EVENT_STATUS_LABELS,
    MEMBER_ROLE_LABELS,
    NOTIFICATION_CHANNEL_LABELS,
    NOTIFICATION_STATUS_LABELS,
    NOTIFICATION_TYPE_LABELS,
    PRIORITY_LABELS,
    READ_STATUS_LABELS,
    REPLY_STATUS_LABELS,
    ROLE_LABELS,
    SUBTASK_STATUS_LABELS,
    TEMPLATE_NOTIFY_TYPE_OPTIONS,
    TASK_STATUS_LABELS,
)
from app.db import SessionLocal, get_db
from app.models import AuditLog, DelayRequest, MailAction, MailEvent, MailScanState, Notification, NotificationRecipient, Task, TaskImportHistory, TaskMember, TaskMilestone, TaskStatusEvent, TaskSubtask, Template, User
from app.schemas import (
    ApiMessage,
    AuditOut,
    DashboardSummary,
    DelayDecisionRequest,
    DelayRequestCreate,
    LoginRequest,
    MailEventDetailOut,
    MailEventOut,
    MailPollStateOut,
    MailBaselineRequest,
    NotificationDetailOut,
    NotificationPreviewOut,
    NotificationRecipientOut,
    NotificationOut,
    RefreshRequest,
    TaskCreate,
    TaskDetailOut,
    TaskImportHistoryOut,
    TaskImportPreviewOut,
    TaskOut,
    TaskStatusUpdate,
    TemplateCreate,
    TemplatePreviewRequest,
    TokenPair,
    UserCreate,
    UserOut,
    UserUpdate,
    RuntimeSettingsUpdate,
)
from app.security import create_token, decode_token, get_current_user, require_admin, verify_password
from app.services.audit import cleanup_system_logs, write_audit
from app.services.collection import collect_state, start_collect
from app.services.delay import apply_delay_decision
from app.services.mail import diagnose_inbox_settings, diagnose_mail_settings, initialize_mail_scan_baseline
from app.services.notifications import create_due_reminders, create_notification_with_recipients, preview_notification_content
from app.services.runtime_settings import load_runtime_settings, runtime_settings_dict, save_runtime_settings
from app.services.templates import sort_templates, template_matches, validate_template_content
from app.timeutils import shanghai_now_naive
from app.services.users import build_default_password_hash, ensure_last_admin_not_removed

router = APIRouter(prefix="/api/v1")
logger = logging.getLogger(__name__)
OPEN_TASK_STATUSES = {"not_started", "in_progress"}
TASK_IMPORT_FIELDS = (
    "sequence",
    "title",
    "content",
    "subtask",
    "responsible_names",
    "deadline",
)
LEGACY_TASK_IMPORT_FIELDS = (
    "title",
    "content",
    "owner_name",
    "participant_names",
    "start_at",
    "end_at",
    "priority",
    "remark",
    "due_remind_days",
    "milestone_names",
    "milestone_datetimes",
    "remind_offsets",
    "subtask_titles",
    "subtask_contents",
    "subtask_assignee_names",
)


def _create_task_created_notifications(db: Session, task_id: int) -> None:
    """为新建任务生成邮件和即时消息通知。

    参数:
    - db: 当前数据库会话，调用方负责提交或回滚事务。
    - task_id: 已经落库的任务编号。

    副作用:
    - 写入通知和接收人记录。
    - 邮件通道会尝试发送邮件，QAX 通道会尝试创建即时消息任务。
    """
    create_notification_with_recipients(db, task_id, "email", "task_created", "")
    create_notification_with_recipients(db, task_id, "qax", "task_created", "")


def _run_task_created_notifications(task_id: int) -> None:
    """在后台会话中发送任务创建通知，避免阻塞创建任务接口。"""
    with SessionLocal() as task_db:
        try:
            _create_task_created_notifications(task_db, task_id)
            task_db.commit()
        except Exception:
            task_db.rollback()
            logger.exception("后台发送任务创建通知失败，任务ID：%s", task_id)


def _enqueue_task_created_notifications(task_id: int) -> None:
    """启动后台线程发送任务创建通知。

    Web 新建任务需要先把任务创建结果返回给前端，邮件和即时消息发送耗时不可阻塞请求。
    """
    thread = threading.Thread(
        target=_run_task_created_notifications,
        args=(task_id,),
        name=f"task-created-notify-{task_id}",
        daemon=True,
    )
    thread.start()


def is_admin_role(role: str) -> bool:
    """判断角色是否属于管理角色。"""
    return role in ADMIN_ROLES


def validate_user_role(role: str) -> None:
    """校验用户角色编码，避免录入无法识别的权限角色。"""
    if role not in ROLE_LABELS:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="用户角色不存在")


def _trim_text(value: str | None) -> str:
    return (value or "").strip()


def _trim_import_match_text(value: object) -> str:
    return "".join(str(value or "").split())


def cleanup_task_scheduled_notifications(db: Session, task_id: int) -> int:
    """删除任务相关的定时提醒通知，并返回清理数量。

    当前系统的定时任务由扫描器按任务配置生成 `due_remind` 通知记录；删除任务时同步清理这些待办提醒，
    可避免列表与通知中心继续出现已经删除任务的定时提醒痕迹。
    """
    notifications = db.query(Notification).filter(Notification.task_id == task_id, Notification.notify_type == "due_remind").all()
    notification_ids = [item.id for item in notifications]
    if not notification_ids:
        return 0
    # 先清理接收人明细，避免保留指向已删除定时提醒的孤儿记录。
    db.query(NotificationRecipient).filter(NotificationRecipient.notification_id.in_(notification_ids)).delete(synchronize_session=False)
    db.query(Notification).filter(Notification.id.in_(notification_ids)).delete(synchronize_session=False)
    return len(notification_ids)


def task_status_text(status: str) -> str:
    """将任务状态编码转换为中文展示文案。"""
    return TASK_STATUS_LABELS.get(status, status)


def task_status_display(task: Task) -> str:
    """生成任务列表与详情页使用的状态展示文本。"""
    text = task_status_text(task.main_status)
    if task.delay_days > 0:
        return f"{text}（延期{task.delay_days}天）"
    return text


def effective_task_delay_days(task: Task, current: datetime) -> int:
    if task.main_status not in OPEN_TASK_STATUSES:
        return 0
    overdue_days = max((current.date() - task.end_at.date()).days, 0)
    return max(int(task.delay_days or 0), overdue_days)


def is_task_due_soon(task: Task, current: datetime, window_days: int = 3) -> bool:
    if task.main_status not in OPEN_TASK_STATUSES or effective_task_delay_days(task, current) > 0:
        return False
    days_to_due = (task.end_at.date() - current.date()).days
    return 0 <= days_to_due <= window_days


def task_import_template_file() -> Path | None:
    api_path = Path(__file__).resolve()
    for root in (api_path.parents[2], api_path.parents[3] if len(api_path.parents) > 3 else None):
        if root is None:
            continue
        template_path = root / "config" / "task-import-template.xlsx"
        if template_path.exists():
            return template_path
    return None


def notification_channel_text(channel: str) -> str:
    return NOTIFICATION_CHANNEL_LABELS.get(channel, channel)


def notification_status_text(status: str) -> str:
    return NOTIFICATION_STATUS_LABELS.get(status, status)


def read_status_text(status: str, channel: str) -> str:
    """按渠道返回“已读/已回复”状态文案。"""
    labels = REPLY_STATUS_LABELS if channel == "email" else READ_STATUS_LABELS
    return labels.get(status, status)


def feedback_label_text(channel: str) -> str:
    """按渠道返回列表与详情页应展示的反馈口径。"""
    return "回复" if channel == "email" else "已读"


def serialize_user(user: User) -> UserOut:
    """将数据库用户对象序列化为接口响应结构。"""
    return UserOut(
        id=user.id,
        username=user.username,
        role=user.role,
        role_text=ROLE_LABELS.get(user.role, user.role),
        name=user.name,
        email=user.email,
        ip_address=user.ip_address,
        is_active=user.is_active,
    )


def _empty_latest_notification(channel: str) -> dict[str, object]:
    return {
        "channel": channel,
        "channel_text": notification_channel_text(channel),
        "notify_type": "",
        "notify_type_text": "",
        "delivery_status": "not_sent",
        "delivery_status_text": "未发送",
        "read_status": "unread",
        "read_status_text": read_status_text("unread", channel),
        "pending_recipients": [],
        "pending_recipient_names": [],
        "pending_recipient_ids": [],
        "sent_at": None,
        "summary": "未发送",
    }


def _latest_notification_summary(
    db: Session,
    task_id: int,
    channel: str,
    *,
    user_id: int | None = None,
) -> dict[str, object]:
    if user_id is None:
        notification = (
            db.query(Notification)
            .filter(Notification.task_id == task_id, Notification.channel == channel)
            .order_by(Notification.id.desc())
            .first()
        )
    else:
        notification = (
            db.query(Notification)
            .join(NotificationRecipient, NotificationRecipient.notification_id == Notification.id)
            .filter(
                Notification.task_id == task_id,
                Notification.channel == channel,
                NotificationRecipient.user_id == user_id,
            )
            .order_by(Notification.id.desc(), NotificationRecipient.id.desc())
            .first()
        )
    if not notification:
        return _empty_latest_notification(channel)

    recipients = db.query(NotificationRecipient).filter(NotificationRecipient.notification_id == notification.id).all()
    summary = _empty_latest_notification(channel)
    summary.update(
        {
            "notification_id": notification.id,
            "notify_type": notification.notify_type,
            "notify_type_text": NOTIFICATION_TYPE_LABELS.get(notification.notify_type, notification.notify_type),
            "sent_at": notification.created_at,
        }
    )

    if user_id is None:
        delivered_count = sum(1 for item in recipients if item.delivery_status == "delivered")
        read_count = sum(1 for item in recipients if item.read_status == "read")
        recipient_total = len(recipients)
        pending_user_ids = [item.user_id for item in recipients if item.read_status != "read"]
        user_names = {
            item.id: item.name
            for item in db.query(User).filter(User.id.in_(pending_user_ids)).all()
        } if pending_user_ids else {}
        pending_recipients = [
            {
                "user_id": item.user_id,
                "name": user_names.get(item.user_id) or f"成员 #{item.user_id}",
            }
            for item in recipients
            if item.read_status != "read"
        ]
        summary.update(
            {
                "delivery_status": notification.status,
                "delivery_status_text": notification_status_text(notification.status),
                "read_status": "read" if read_count > 0 else "unread",
                "read_status_text": f"{read_count}/{recipient_total} {feedback_label_text(channel)}" if recipient_total else read_status_text("unread", channel),
                "pending_recipients": pending_recipients,
                "pending_recipient_names": [item["name"] for item in pending_recipients],
                "pending_recipient_ids": [item["user_id"] for item in pending_recipients],
                "summary": (
                    f"已送达 {delivered_count}/{recipient_total}"
                    if recipient_total
                    else notification_status_text(notification.status)
                ),
            }
        )
        return summary

    recipient = next((item for item in recipients if item.user_id == user_id), None)
    if not recipient:
        return summary
    summary.update(
        {
            "delivery_status": recipient.delivery_status,
            "delivery_status_text": notification_status_text(recipient.delivery_status),
            "read_status": recipient.read_status,
            "read_status_text": read_status_text(recipient.read_status, channel),
            "summary": (
                read_status_text(recipient.read_status, channel)
                if recipient.delivery_status == "delivered"
                else notification_status_text(recipient.delivery_status)
            ),
        }
    )
    return summary


def _task_latest_notifications(db: Session, task_id: int, *, user_id: int | None = None) -> dict[str, dict[str, object]]:
    return {
        "email": _latest_notification_summary(db, task_id, "email", user_id=user_id),
        "qax": _latest_notification_summary(db, task_id, "qax", user_id=user_id),
    }


def serialize_task(task: Task, db: Session) -> TaskOut:
    """聚合任务成员和通知统计，生成任务列表项。"""
    members = db.query(TaskMember).filter(TaskMember.task_id == task.id).all()
    owner_name = next((item.user.name for item in members if item.member_role == "owner" and item.user), "")
    creator = db.query(User).filter(User.id == task.created_by).first() if task.created_by else None
    responsible_names = [item.user.name for item in members if item.user]
    participant_count = sum(1 for item in members if item.member_role == "participant")
    recipients = (
        db.query(NotificationRecipient)
        .join(Notification, NotificationRecipient.notification_id == Notification.id)
        .filter(Notification.task_id == task.id)
        .all()
    )
    delivered_count = sum(1 for item in recipients if item.delivery_status == "delivered")
    subtask_status_summary = []
    subtasks = list(task.subtasks or [])
    for status in ("pending", "in_progress", "done", "canceled"):
        count = sum(1 for item in subtasks if item.status == status)
        if count <= 0:
            continue
        subtask_status_summary.append(
            {
                "status": status,
                "label": SUBTASK_STATUS_LABELS.get(status, status),
                "count": count,
            }
        )
    return TaskOut(
        id=task.id,
        title=task.title,
        content=task.content,
        remark=task.remark,
        priority=task.priority,
        main_status=task.main_status,
        delay_days=task.delay_days,
        state_locked=task.state_locked,
        due_remind_days=task.due_remind_days,
        start_at=task.start_at,
        end_at=task.end_at,
        planned_minutes=task.planned_minutes,
        actual_minutes=task.actual_minutes,
        completed_at=task.completed_at,
        status_text=task_status_display(task),
        priority_text=PRIORITY_LABELS.get(task.priority, task.priority),
        owner_name=owner_name,
        creator_name=creator.name if creator else "",
        responsible_names=responsible_names,
        participant_count=participant_count,
        notification_total=len(recipients),
        delivered_count=delivered_count,
        completed_member_count=1 if task.main_status == "done" else 0,
        subtask_count=len(subtasks),
        subtask_status_summary=subtask_status_summary,
        latest_notifications=_task_latest_notifications(db, task.id),
        created_at=task.created_at,
    )


def extract_remind_focus(content_snapshot: str) -> str:
    """从通知正文快照中提取“当前提醒重点”，兼容历史已落库通知正文。"""
    if not content_snapshot:
        return ""
    marker = "当前提醒重点："
    for line in content_snapshot.replace("\r\n", "\n").split("\n"):
        if line.startswith(marker):
            return line.replace(marker, "", 1).strip()
    return ""


def notification_scene_text(notify_type: str, remind_focus: str) -> str:
    """将提醒场景转换为更贴近业务的中文文案，方便列表和详情快速识别。"""
    if notify_type != "manual_remind":
        return NOTIFICATION_TYPE_LABELS.get(notify_type, notify_type)
    if remind_focus.startswith("请优先跟进子任务"):
        return "子任务提醒"
    if remind_focus.startswith("请围绕里程碑"):
        return "里程碑提醒"
    return "任务提醒"


def serialize_notification(notification: Notification, db: Session) -> NotificationOut:
    """将通知主记录扩展为包含统计信息的列表项。"""
    recipients = db.query(NotificationRecipient).filter(NotificationRecipient.notification_id == notification.id).all()
    task_title = ""
    if notification.task_id:
        task = db.query(Task).filter(Task.id == notification.task_id).first()
        if task:
            task_title = task.title
    remind_focus = extract_remind_focus(notification.content_snapshot)
    return NotificationOut(
        id=notification.id,
        task_id=notification.task_id,
        task_title=task_title,
        channel=notification.channel,
        notify_type=notification.notify_type,
        notify_type_text=NOTIFICATION_TYPE_LABELS.get(notification.notify_type, notification.notify_type),
        notify_scene_text=notification_scene_text(notification.notify_type, remind_focus),
        feedback_label=feedback_label_text(notification.channel),
        status=notification.status,
        channel_text=notification_channel_text(notification.channel),
        status_text=notification_status_text(notification.status),
        recipient_total=len(recipients),
        delivered_count=sum(1 for item in recipients if item.delivery_status == "delivered"),
        read_count=sum(1 for item in recipients if item.read_status == "read"),
        retry_total=sum(item.retry_count for item in recipients),
        last_error=next((item.last_error for item in recipients if item.last_error), ""),
        remind_focus=remind_focus,
        created_at=notification.created_at,
    )


def serialize_notification_recipient(recipient: NotificationRecipient, db: Session) -> NotificationRecipientOut:
    """序列化通知接收人明细。"""
    user = db.query(User).filter(User.id == recipient.user_id).first()
    notification = db.query(Notification).filter(Notification.id == recipient.notification_id).first()
    channel = notification.channel if notification else "email"
    return NotificationRecipientOut(
        user_id=recipient.user_id,
        name=user.name if user else "",
        email=user.email if user else "",
        recipient_role=recipient.recipient_role,
        recipient_role_text=MEMBER_ROLE_LABELS.get(recipient.recipient_role, recipient.recipient_role),
        delivery_status=recipient.delivery_status,
        delivery_status_text=notification_status_text(recipient.delivery_status),
        read_status=recipient.read_status,
        read_status_text=read_status_text(recipient.read_status, channel),
        feedback_label=feedback_label_text(channel),
        retry_count=recipient.retry_count,
        content_snapshot=recipient.content_snapshot,
        last_error=recipient.last_error,
    )


def serialize_notification_detail(notification: Notification, db: Session) -> NotificationDetailOut:
    """构造通知详情页所需的完整响应。"""
    base = serialize_notification(notification, db)
    recipients = db.query(NotificationRecipient).filter(NotificationRecipient.notification_id == notification.id).all()
    return NotificationDetailOut(
        **base.model_dump(),
        content_snapshot=notification.content_snapshot,
        recipients=[serialize_notification_recipient(item, db) for item in recipients],
    )


def serialize_task_import_history(history: TaskImportHistory, db: Session) -> TaskImportHistoryOut:
    """序列化任务导入历史。"""
    operator = db.query(User).filter(User.id == history.operator_id).first()
    try:
        summary = json.loads(history.summary_json or "{}")
    except json.JSONDecodeError:
        summary = {}
    return TaskImportHistoryOut(
        id=history.id,
        filename=history.filename,
        operator_name=operator.name if operator else "",
        total_rows=history.total_rows,
        success_count=history.success_count,
        failure_count=history.failure_count,
        overlap_count=history.overlap_count,
        confirmed_duplicate=history.confirmed_duplicate,
        overlap_samples=summary.get("overlap_samples", [])[:5],
        failure_samples=summary.get("failures", [])[:5],
        created_at=history.created_at,
    )


def serialize_mail_event(mail_event: MailEvent, db: Session) -> MailEventOut:
    """序列化邮件事件，并补齐匹配模板与动作结果。"""
    template = db.query(Template).filter(Template.id == mail_event.resolved_template_id).first() if mail_event.resolved_template_id else None
    action = db.query(MailAction).filter(MailAction.mail_event_id == mail_event.id).order_by(MailAction.id.desc()).first()
    task = db.query(Task).filter(Task.id == action.target_task_id).first() if action and action.target_task_id else None
    notify_type = template.notify_type if template else ""
    return MailEventOut(
        id=mail_event.id,
        message_id=mail_event.message_id,
        from_addr=mail_event.from_addr,
        subject=mail_event.subject,
        body_digest=mail_event.body_digest,
        process_status=mail_event.process_status,
        process_status_text=MAIL_EVENT_STATUS_LABELS.get(mail_event.process_status, mail_event.process_status),
        template_name=template.name if template else "",
        notify_type=notify_type,
        notify_type_text=NOTIFICATION_TYPE_LABELS.get(notify_type, notify_type),
        task_id=action.target_task_id if action else None,
        task_title=task.title if task else "",
        action_type=action.action_type if action else "",
        action_status=action.action_status if action else "",
        action_status_text=MAIL_ACTION_STATUS_LABELS.get(action.action_status, action.action_status) if action else "",
        action_result_json=action.action_result_json if action else "",
        created_at=mail_event.created_at,
    )


def serialize_mail_event_detail(mail_event: MailEvent, db: Session) -> MailEventDetailOut:
    """序列化邮件详情，补充模板正文快照。"""
    base = serialize_mail_event(mail_event, db)
    template = db.query(Template).filter(Template.id == mail_event.resolved_template_id).first() if mail_event.resolved_template_id else None
    return MailEventDetailOut(
        **base.model_dump(),
        template_id=template.id if template else None,
        template_kind=template.template_kind if template else "",
        content=template.content if template else "",
        original_body=mail_event.original_body,
    )


def serialize_mail_poll_state(db: Session) -> MailPollStateOut:
    """读取自动收件配置和最近扫描时间，供前端展示倒计时。"""
    state = db.query(MailScanState).filter(MailScanState.id == 1).first()
    runtime = load_runtime_settings()
    interval_seconds = max(runtime.mail_auto_poll_interval_seconds, 30)
    last_scan_at = state.last_scan_at if state else None
    next_poll_at = None
    if runtime.mail_auto_poll_enabled and last_scan_at:
        next_poll_at = last_scan_at + timedelta(seconds=interval_seconds)
    protocol = settings.mail_inbox_protocol if settings.mail_inbox_protocol in {"imap", "pop3"} else "imap"
    return MailPollStateOut(
        inbox_protocol=protocol,
        inbox_protocol_text="POP3" if protocol == "pop3" else "IMAP",
        auto_poll_enabled=runtime.mail_auto_poll_enabled,
        interval_seconds=interval_seconds,
        last_scan_at=last_scan_at,
        next_poll_at=next_poll_at,
        baseline_started_at=state.baseline_started_at if state else None,
        qax_auto_collect_enabled=runtime.qax_auto_collect_enabled,
        qax_interval_seconds=max(runtime.qax_auto_collect_interval_seconds, 30),
        qax_browser_visible=runtime.qax_browser_visible,
    )


def ensure_notification_access(notification: Notification, current_user: User, db: Session) -> None:
    """校验当前用户是否允许查看某条通知详情。"""
    if is_admin_role(current_user.role):
        return
    if notification.task_id is None:
        raise HTTPException(status_code=403, detail="无权访问该通知")
    membership = db.query(TaskMember).filter(TaskMember.task_id == notification.task_id, TaskMember.user_id == current_user.id).first()
    if not membership:
        raise HTTPException(status_code=403, detail="无权访问该通知")


def _create_task_record(payload: TaskCreate, current_user: User, db: Session, source: str = "web") -> Task:
    """按统一规则落库任务、成员、里程碑、状态流水与初始通知。

    Web 创建入口以服务端当前时间作为任务开始时间，避免旧前端或手工请求传入历史开始时间后，
    让“任务是否开始”的展示口径重新退回到按时间判断；导入入口仍保留文件中的开始时间。
    """
    actual_start_at = shanghai_now_naive() if source == "web" else payload.start_at
    if actual_start_at > payload.end_at:
        raise HTTPException(status_code=400, detail="开始时间不能晚于结束时间")

    initial_status = "not_started"
    task = Task(
        title=payload.title,
        content=payload.content,
        priority=payload.priority,
        remark=payload.remark,
        start_at=actual_start_at,
        end_at=payload.end_at,
        due_remind_days=max(payload.due_remind_days, 0),
        planned_minutes=int((payload.end_at - actual_start_at).total_seconds() // 60),
        main_status=initial_status,
        completed_at=None,
        created_by=current_user.id,
    )
    db.add(task)
    db.flush()

    # 负责人也要作为成员写入，便于成员视图与通知逻辑统一复用同一张关联表。
    member_ids = {payload.owner_id, *payload.participant_ids}
    for user_id in member_ids:
        member_role = "owner" if user_id == payload.owner_id else "participant"
        db.add(TaskMember(task_id=task.id, user_id=user_id, member_role=member_role))

    for milestone in payload.milestones:
        if milestone.planned_at < actual_start_at or milestone.planned_at > payload.end_at:
            raise HTTPException(status_code=400, detail="里程碑时间必须位于任务时间范围内")
        db.add(
            TaskMilestone(
                task_id=task.id,
                name=milestone.name,
                planned_at=milestone.planned_at,
                remind_offsets=",".join(str(item) for item in milestone.remind_offsets),
                sort_order=milestone.sort_order,
            )
        )

    for subtask in payload.subtasks:
        if subtask.assignee_id not in member_ids:
            raise HTTPException(status_code=400, detail="子任务执行人必须从主任务参与成员中选择")
        db.add(
            TaskSubtask(
                task_id=task.id,
                title=subtask.title,
                content=subtask.content,
                assignee_id=subtask.assignee_id,
                sort_order=subtask.sort_order,
                status=subtask.status or "pending",
            )
        )

    db.add(TaskStatusEvent(task_id=task.id, from_status="", to_status=initial_status, source=source, remark="创建任务，等待成员确认已读", operator_id=current_user.id))
    write_audit(
        db,
        current_user.id,
        "CREATE_TASK",
        "Task",
        task.id,
        {},
        {"title": task.title, "source": source},
        module_name="api.task",
        message=f"创建任务《{task.title}》",
        detail={
            "source": source,
            # 这里直接复用入参中的负责人 ID，避免依赖当前作用域之外的临时变量。
            "owner_id": payload.owner_id,
            "participant_ids": [item.user_id for item in task.members if item.member_role != "owner"],
            "subtask_total": len(payload.subtasks),
            "milestone_total": len(payload.milestones),
        },
    )
    return task


def _normalize_import_datetime(value: object, field_label: str) -> datetime:
    """将 Excel 单元格值归一化为 datetime。"""
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    text = str(value or "").strip()
    if not text:
        raise ValueError(f"{field_label}不能为空")
    try:
        return datetime.fromisoformat(text)
    except ValueError as exc:
        raise ValueError(f"{field_label}格式无效，请使用 YYYY-MM-DDTHH:MM:SS") from exc


def _split_import_usernames(value: object) -> list[str]:
    """解析 Excel 中的姓名列表。"""
    text = str(value or "").replace("，", ",").replace("；", ",").replace(" ", "").strip(",")
    if not text:
        return []
    return [item for item in text.split(",") if item]


def _find_active_user_by_name(db: Session, name: str, field_label: str) -> User:
    """按姓名查找启用中的系统用户，并阻止重名误导入。"""
    user_name = _trim_import_match_text(name)
    if not user_name:
        raise ValueError(f"{field_label}不能为空")
    matches = [
        user
        for user in db.query(User).filter(User.is_active.is_(True)).all()
        if _trim_import_match_text(user.name) == user_name
    ]
    if not matches:
        raise ValueError(f"{field_label}不存在或已禁用：{user_name}")
    if len(matches) > 1:
        raise ValueError(f"{field_label}存在重名用户，请先处理重名：{user_name}")
    return matches[0]


def _build_import_row_signature(row_data: dict[str, object]) -> str:
    """为导入行生成稳定签名，用于识别高重叠重复导入。"""
    signature_payload = {
        "title": str(row_data.get("title") or "").strip(),
        "content": str(row_data.get("content") or "").strip(),
        "owner_name": str(row_data.get("owner_name") or "").strip(),
        "participant_names": sorted(_split_import_usernames(row_data.get("participant_names"))),
        "start_at": str(row_data.get("start_at") or "").strip(),
        "end_at": str(row_data.get("end_at") or "").strip(),
        "priority": str(row_data.get("priority") or "").strip().lower(),
        "subtask_titles": str(row_data.get("subtask_titles") or "").strip(),
        "subtask_assignee_names": str(row_data.get("subtask_assignee_names") or "").strip(),
    }
    return hashlib.sha256(json.dumps(signature_payload, ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest()


def _build_import_row_sample(row_number: int, row_data: dict[str, object], signature: str) -> dict[str, object]:
    """构造导入行摘要，用于历史回看和重复对比展示。"""
    return {
        "signature": signature,
        "row_number": row_number,
        "title": str(row_data.get("title") or "").strip(),
        "owner_name": str(row_data.get("owner_name") or "").strip(),
        "end_at": str(row_data.get("end_at") or "").strip(),
    }


def _collect_import_rows(rows: list[tuple[object, ...]]) -> tuple[list[tuple[int, dict[str, object]]], list[str], list[dict[str, object]]]:
    """抽取有效导入行，并预先生成行签名。"""
    collected_rows: list[tuple[int, dict[str, object]]] = []
    signatures: list[str] = []
    row_samples: list[dict[str, object]] = []
    for index, row in enumerate(rows[1:], start=2):
        if not any(cell not in (None, "") for cell in row):
            continue
        row_data = {field: row[position] if position < len(row) else None for position, field in enumerate(LEGACY_TASK_IMPORT_FIELDS)}
        signature = _build_import_row_signature(row_data)
        collected_rows.append((index, row_data))
        signatures.append(signature)
        row_samples.append(_build_import_row_sample(index, row_data, signature))
    return collected_rows, signatures, row_samples


def _detect_import_overlap(db: Session, signatures: list[str], row_samples: list[dict[str, object]]) -> tuple[int, float, list[dict[str, object]]]:
    """检测当前导入与近几次导入历史的最大重叠规模。"""
    current_signatures = {item for item in signatures if item}
    current_sample_map = {item.get("signature"): item for item in row_samples if item.get("signature")}
    if not current_signatures:
        return 0, 0, []
    best_history: TaskImportHistory | None = None
    best_overlap_signatures: set[str] = set()
    histories = db.query(TaskImportHistory).order_by(TaskImportHistory.id.desc()).limit(20).all()
    for history in histories:
        history_signatures = set(json.loads(history.row_signatures_json or "[]"))
        overlap_signatures = current_signatures & history_signatures
        if len(overlap_signatures) > len(best_overlap_signatures):
            best_history = history
            best_overlap_signatures = overlap_signatures
    overlap_count = len(best_overlap_signatures)
    overlap_rate = overlap_count / len(current_signatures) if current_signatures else 0
    overlap_samples: list[dict[str, object]] = []
    if best_history and best_overlap_signatures:
        try:
            history_summary = json.loads(best_history.summary_json or "{}")
        except json.JSONDecodeError:
            history_summary = {}
        history_rows = {
            item.get("signature"): item
            for item in history_summary.get("rows", [])
            if isinstance(item, dict) and item.get("signature")
        }
        for signature in list(best_overlap_signatures)[:5]:
            current_row = current_sample_map.get(signature, {})
            history_row = history_rows.get(signature, {})
            overlap_samples.append(
                {
                    "title": current_row.get("title") or history_row.get("title") or "-",
                    "owner_name": current_row.get("owner_name") or history_row.get("owner_name") or "",
                    "end_at": current_row.get("end_at") or history_row.get("end_at") or "",
                    "current_row_number": current_row.get("row_number") or history_row.get("row_number"),
                    "history_filename": best_history.filename,
                    "history_created_at": best_history.created_at.isoformat() if best_history.created_at else "",
                }
            )
    return overlap_count, overlap_rate, overlap_samples


def _parse_import_milestones(row_data: dict[str, object]) -> list[dict]:
    """解析导入行中的里程碑字段。"""
    milestone_names = [item.strip() for item in str(row_data.get("milestone_names") or "").split("|") if item.strip()]
    milestone_datetimes = [item.strip() for item in str(row_data.get("milestone_datetimes") or "").split("|") if item.strip()]
    remind_groups = [item.strip() for item in str(row_data.get("remind_offsets") or "").split("|") if item.strip()]

    if not milestone_names and not milestone_datetimes and not remind_groups:
        return []
    if len(milestone_names) != len(milestone_datetimes):
        raise ValueError("里程碑名称与里程碑时间数量不一致")
    if remind_groups and len(remind_groups) != len(milestone_names):
        raise ValueError("里程碑提醒天数数量与里程碑名称数量不一致")

    milestones: list[dict] = []
    for index, name in enumerate(milestone_names):
        remind_values = remind_groups[index] if remind_groups else "1"
        offsets = [int(item.strip()) for item in remind_values.split(",") if item.strip()]
        milestones.append(
            {
                "name": name,
                "planned_at": _normalize_import_datetime(milestone_datetimes[index], f"第 {index + 1} 个里程碑时间"),
                "remind_offsets": offsets or [1],
                "sort_order": index,
            }
        )
    return milestones


def _parse_import_subtasks(row_data: dict[str, object], db: Session, member_ids: set[int]) -> list[dict]:
    """解析导入行中的子任务字段。"""
    subtask_titles = [item.strip() for item in str(row_data.get("subtask_titles") or "").split("|") if item.strip()]
    subtask_contents = [item.strip() for item in str(row_data.get("subtask_contents") or "").split("|")] if row_data.get("subtask_contents") else []
    subtask_assignees = [item.strip() for item in str(row_data.get("subtask_assignee_names") or "").split("|") if item.strip()]

    if not subtask_titles and not subtask_assignees and not any(item for item in subtask_contents):
        return []
    if len(subtask_titles) != len(subtask_assignees):
        raise ValueError("子任务标题与子任务执行人数量不一致")
    if subtask_contents and len(subtask_contents) not in (0, len(subtask_titles)):
        raise ValueError("子任务内容数量与子任务标题数量不一致")

    subtasks: list[dict] = []
    for index, title in enumerate(subtask_titles):
        assignee_name = subtask_assignees[index]
        assignee = _find_active_user_by_name(db, assignee_name, "子任务执行人")
        if assignee.id not in member_ids:
            raise ValueError(f"子任务执行人必须属于主任务成员：{assignee_name}")
        subtasks.append(
            {
                "title": title,
                "content": subtask_contents[index] if index < len(subtask_contents) else "",
                "assignee_id": assignee.id,
                "sort_order": index,
            }
        )
    return subtasks


def _build_task_create_from_import_row(row_data: dict[str, object], db: Session) -> TaskCreate:
    """将 Excel 行数据转换成任务创建模型。"""
    owner = _find_active_user_by_name(db, str(row_data.get("owner_name") or "").strip(), "负责人")

    participant_usernames = _split_import_usernames(row_data.get("participant_names"))
    participant_ids: list[int] = []
    for user_name in participant_usernames:
        user = _find_active_user_by_name(db, user_name, "参与人员")
        if user.id != owner.id and user.id not in participant_ids:
            participant_ids.append(user.id)
    member_ids = {owner.id, *participant_ids}

    priority = str(row_data.get("priority") or "").strip().lower()
    if priority not in PRIORITY_LABELS:
        raise ValueError("优先级仅支持 high、medium、low")

    due_remind_days_value = row_data.get("due_remind_days")
    due_remind_days = 0 if due_remind_days_value in (None, "") else int(due_remind_days_value)
    if due_remind_days < 0:
        raise ValueError("到期前提醒天数不能小于 0")

    payload = {
        "title": str(row_data.get("title") or "").strip(),
        "content": str(row_data.get("content") or "").strip(),
        "owner_id": owner.id,
        "participant_ids": participant_ids,
        "start_at": _normalize_import_datetime(row_data.get("start_at"), "开始时间"),
        "end_at": _normalize_import_datetime(row_data.get("end_at"), "结束时间"),
        "due_remind_days": due_remind_days,
        "priority": priority,
        "remark": str(row_data.get("remark") or "").strip(),
        "milestones": _parse_import_milestones(row_data),
        "subtasks": _parse_import_subtasks(row_data, db, member_ids),
    }
    if not payload["title"]:
        raise ValueError("任务标题不能为空")
    if not payload["content"]:
        raise ValueError("任务内容不能为空")
    return TaskCreate(**payload)


def _split_simple_import_names(value: object) -> list[str]:
    text = str(value or "")
    for marker in ("、", "，", "；", ";", "/", "|"):
        text = text.replace(marker, ",")
    text = _trim_import_match_text(text).strip(",")
    if not text:
        return []
    names: list[str] = []
    for item in text.split(","):
        normalized = item.strip()
        if normalized and normalized not in names:
            names.append(normalized)
    return names


def _normalize_simple_import_deadline(value: object) -> datetime:
    deadline = _normalize_import_datetime(value, "截止日期")
    if deadline.hour == 0 and deadline.minute == 0 and deadline.second == 0:
        return deadline.replace(hour=23, minute=59, second=0, microsecond=0)
    return deadline


def _collect_simple_import_rows(rows: list[tuple[object, ...]]) -> tuple[list[tuple[int, dict[str, object]]], list[str], list[dict[str, object]]]:
    collected: dict[str, dict[str, object]] = {}
    for index, row in enumerate(rows[1:], start=2):
        if not any(cell not in (None, "") for cell in row):
            continue
        first_cell = str(row[0] or "").strip()
        if first_cell.startswith("补充说明"):
            break
        row_data = {field: row[position] if position < len(row) else None for position, field in enumerate(TASK_IMPORT_FIELDS)}
        title = str(row_data.get("title") or "").strip()
        if not title or title == "……":
            continue
        key = str(row_data.get("sequence") or "").strip() or f"row-{index}"
        grouped = collected.setdefault(
            key,
            {
                "sequence": key,
                "title": title,
                "content": str(row_data.get("content") or "").strip(),
                "responsible_names": [],
                "deadline": row_data.get("deadline"),
                "subtasks": [],
                "row_number": index,
            },
        )
        if not grouped.get("content") and row_data.get("content"):
            grouped["content"] = str(row_data.get("content") or "").strip()
        if not grouped.get("deadline") and row_data.get("deadline"):
            grouped["deadline"] = row_data.get("deadline")
        for name in _split_simple_import_names(row_data.get("responsible_names")):
            if name not in grouped["responsible_names"]:
                grouped["responsible_names"].append(name)
        subtask_title = str(row_data.get("subtask") or "").strip()
        if subtask_title:
            for assignee_name in _split_simple_import_names(row_data.get("responsible_names")):
                grouped["subtasks"].append({"title": subtask_title, "assignee_name": assignee_name})

    import_rows: list[tuple[int, dict[str, object]]] = []
    row_signatures: list[str] = []
    row_samples: list[dict[str, object]] = []
    for grouped in collected.values():
        row_data = {
            "sequence": grouped["sequence"],
            "title": grouped["title"],
            "content": grouped["content"],
            "responsible_names": "、".join(grouped["responsible_names"]),
            "deadline": grouped["deadline"],
            "subtasks": grouped["subtasks"],
        }
        # Excel 读取到的截止日期可能是 datetime；先标准化再生成签名，避免导入预检直接抛出序列化异常。
        normalized_signature_payload = {
            **row_data,
            "deadline": row_data["deadline"].isoformat() if isinstance(row_data["deadline"], datetime) else row_data["deadline"],
        }
        signature = hashlib.sha256(
            json.dumps(normalized_signature_payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
        ).hexdigest()
        row_number = int(grouped["row_number"])
        import_rows.append((row_number, row_data))
        row_signatures.append(signature)
        row_samples.append(
            {
                "signature": signature,
                "row_number": row_number,
                "title": row_data["title"],
                "owner_name": row_data["responsible_names"],
                "end_at": str(row_data["deadline"] or ""),
            }
        )
    return import_rows, row_signatures, row_samples


def _build_task_create_from_simple_import_row(row_data: dict[str, object], db: Session) -> TaskCreate:
    responsible_names = _split_simple_import_names(row_data.get("responsible_names"))
    if not responsible_names:
        raise ValueError("负责人不能为空")
    responsible_users = [_find_active_user_by_name(db, item, "负责人") for item in responsible_names]
    owner = responsible_users[0]
    participant_ids = [item.id for item in responsible_users[1:]]
    member_ids = {owner.id, *participant_ids}
    end_at = _normalize_simple_import_deadline(row_data.get("deadline"))
    now = shanghai_now_naive().replace(second=0, microsecond=0)
    start_at = now if now < end_at else end_at.replace(hour=0, minute=0, second=0, microsecond=0)
    if start_at >= end_at:
        start_at = end_at - timedelta(hours=1)

    subtasks: list[dict[str, object]] = []
    for index, item in enumerate(row_data.get("subtasks") or []):
        assignee = _find_active_user_by_name(db, item.get("assignee_name"), "子任务负责人")
        if assignee.id not in member_ids:
            raise ValueError(f"子任务负责人不在任务负责人范围内：{item.get('assignee_name')}")
        subtasks.append(
            {
                "title": str(item.get("title") or "").strip(),
                "content": "",
                "assignee_id": assignee.id,
                "sort_order": index,
                "status": "pending",
            }
        )

    return TaskCreate(
        title=str(row_data.get("title") or "").strip(),
        content=str(row_data.get("content") or "").strip(),
        owner_id=owner.id,
        participant_ids=participant_ids,
        start_at=start_at,
        end_at=end_at,
        due_remind_days=0,
        priority="medium",
        remark="",
        milestones=[],
        subtasks=subtasks,
    )


def _record_task_import_history(
    db: Session,
    filename: str,
    current_user: User,
    row_signatures: list[str],
    row_samples: list[dict[str, object]],
    success_count: int,
    failure_count: int,
    overlap_count: int,
    confirmed_duplicate: bool,
    failures: list[dict[str, object]],
    overlap_samples: list[dict[str, object]],
    workbook_bytes: bytes,
) -> TaskImportHistory:
    """记录任务导入批次结果，供后续历史查询和重复检测。"""
    history = TaskImportHistory(
        filename=filename,
        operator_id=current_user.id,
        file_hash=hashlib.sha256(workbook_bytes).hexdigest(),
        total_rows=len(row_signatures),
        success_count=success_count,
        failure_count=failure_count,
        overlap_count=overlap_count,
        confirmed_duplicate=confirmed_duplicate,
        row_signatures_json=json.dumps(row_signatures, ensure_ascii=False),
        summary_json=json.dumps(
            {
                "failures": failures[:20],
                "rows": row_samples,
                "overlap_samples": overlap_samples[:10],
            },
            ensure_ascii=False,
        ),
    )
    db.add(history)
    db.flush()
    return history


@router.post("/auth/login", response_model=TokenPair)
def login(payload: LoginRequest, db: Session = Depends(get_db)) -> TokenPair:
    """用户名密码登录，并返回访问令牌与刷新令牌。"""
    user = db.query(User).filter(User.username == payload.username, User.is_active.is_(True)).first()
    if not user or not verify_password(payload.password, user.password_hash):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="用户名或密码错误")
    return TokenPair(
        access_token=create_token(str(user.id), "access", settings.access_token_expire_minutes),
        refresh_token=create_token(str(user.id), "refresh", settings.refresh_token_expire_minutes),
    )


@router.post("/auth/refresh", response_model=TokenPair)
def refresh_tokens(payload: RefreshRequest) -> TokenPair:
    try:
        decoded = decode_token(payload.refresh_token)
    except jwt.PyJWTError as exc:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="刷新令牌无效") from exc
    if decoded.get("type") != "refresh":
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="刷新令牌类型错误")
    user_id = str(decoded["sub"])
    return TokenPair(
        access_token=create_token(user_id, "access", settings.access_token_expire_minutes),
        refresh_token=create_token(user_id, "refresh", settings.refresh_token_expire_minutes),
    )


@router.post("/auth/logout", response_model=ApiMessage)
def logout(_: User = Depends(get_current_user)) -> ApiMessage:
    return ApiMessage(message="已退出登录")


@router.get("/auth/me", response_model=UserOut)
def me(current_user: User = Depends(get_current_user)) -> UserOut:
    return serialize_user(current_user)


@router.get("/admin/users", response_model=list[UserOut])
def list_users(_: User = Depends(require_admin), db: Session = Depends(get_db)) -> list[UserOut]:
    return [
        serialize_user(user)
        for user in db.query(User).filter(~User.username.like("deleted__%")).order_by(User.id.asc()).all()
    ]


@router.post("/admin/users", response_model=UserOut)
def create_user(payload: UserCreate, current_user: User = Depends(require_admin), db: Session = Depends(get_db)) -> UserOut:
    validate_user_role(payload.role)
    username = _trim_text(payload.username)
    name = _trim_text(payload.name)
    email = _trim_text(payload.email)
    ip_address = _trim_text(payload.ip_address)
    user = User(
        username=username,
        password_hash=build_default_password_hash(settings.default_password),
        role=payload.role,
        name=name,
        email=email,
        ip_address=ip_address,
        is_active=True,
    )
    db.add(user)
    db.flush()
    write_audit(
        db,
        current_user.id,
        "CREATE_USER",
        "User",
        user.id,
        {},
        {"username": user.username, "role": user.role},
        module_name="api.user",
        message=f"创建用户「{user.name}」",
        detail={"username": user.username, "role": user.role, "email": user.email, "ip_address": user.ip_address},
    )
    db.commit()
    db.refresh(user)
    return serialize_user(user)


@router.put("/admin/users/{user_id}", response_model=UserOut)
def update_user(user_id: int, payload: UserUpdate, current_user: User = Depends(require_admin), db: Session = Depends(get_db)) -> UserOut:
    validate_user_role(payload.role)
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    ensure_last_admin_not_removed(db, user, payload.role, payload.is_active)
    before = {"role": user.role, "is_active": user.is_active, "email": user.email, "ip_address": user.ip_address}
    user.role = payload.role
    user.name = _trim_text(payload.name)
    user.email = _trim_text(payload.email)
    user.ip_address = _trim_text(payload.ip_address)
    user.is_active = payload.is_active
    write_audit(
        db,
        current_user.id,
        "UPDATE_USER",
        "User",
        user.id,
        before,
        {"role": user.role, "is_active": user.is_active},
        module_name="api.user",
        message=f"更新用户「{user.name}」信息",
        detail={"email": user.email, "ip_address": user.ip_address, "is_active": user.is_active},
    )
    db.commit()
    db.refresh(user)
    return serialize_user(user)


@router.delete("/admin/users/{user_id}", response_model=ApiMessage)
def delete_user(user_id: int, current_user: User = Depends(require_admin), db: Session = Depends(get_db)) -> ApiMessage:
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    ensure_last_admin_not_removed(db, user, new_active=False)
    before = {"username": user.username, "role": user.role, "email": user.email, "ip_address": user.ip_address}
    user.username = f"deleted__{user.id}"
    user.email = f"deleted__{user.id}@deleted.local"
    user.ip_address = f"deleted__{user.id}"
    user.name = f"{user.name}(已删除)"
    user.role = "member"
    user.is_active = False
    write_audit(
        db,
        current_user.id,
        "DELETE_USER",
        "User",
        user_id,
        before,
        {"deleted": True},
        module_name="api.user",
        message=f"删除用户 {user.name}",
        detail=before,
    )
    db.commit()
    return ApiMessage(message="用户已删除")


@router.get("/dashboard/summary", response_model=DashboardSummary)
def dashboard_summary(_: User = Depends(require_admin), db: Session = Depends(get_db)) -> DashboardSummary:
    now = shanghai_now_naive()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    tasks = db.query(Task).filter(Task.deleted_at.is_(None)).all()
    notifications = db.query(Notification).all()
    recipients = db.query(NotificationRecipient).all()
    delay_requests = db.query(DelayRequest).all()
    mail_events = db.query(MailEvent).all()
    status_events = db.query(TaskStatusEvent).all()
    email_notifications = [item for item in notifications if item.channel == "email"]
    qax_ids = {item.id for item in notifications if item.channel == "qax"}
    qax_recipients = [item for item in recipients if item.notification_id in qax_ids]
    email_success = sum(1 for item in email_notifications if item.status in ("sent", "delivered"))
    pending_total = sum(1 for item in tasks if item.main_status == "not_started")
    in_progress_total = sum(1 for item in tasks if item.main_status == "in_progress")
    done_total = sum(1 for item in tasks if item.main_status == "done")
    canceled_total = sum(1 for item in tasks if item.main_status == "canceled")
    delayed_total = sum(1 for item in tasks if effective_task_delay_days(item, now) > 0)
    due_soon_total = sum(1 for item in tasks if is_task_due_soon(item, now))
    pending_delay_requests = sum(1 for item in delay_requests if item.approval_status == "PENDING")
    failed_recipients = sum(1 for item in recipients if item.delivery_status == "failed")
    failed_mail_events = sum(1 for item in mail_events if item.process_status == "FAILED")
    mail_failure_total = failed_recipients + failed_mail_events
    completion_rate = round(done_total * 100 / len(tasks), 2) if tasks else 0.0
    healthy_task_base = max(len(tasks) - canceled_total, 1)
    healthy_task_rate = round((healthy_task_base - delayed_total) * 100 / healthy_task_base, 2) if tasks else 0.0
    email_success_rate = round(email_success * 100 / len(email_notifications), 2) if email_notifications else 0.0
    qax_delivery_rate = round(sum(1 for item in qax_recipients if item.delivery_status == "delivered") * 100 / len(qax_recipients), 2) if qax_recipients else 0.0
    qax_read_rate = round(sum(1 for item in qax_recipients if item.read_status == "read") * 100 / len(qax_recipients), 2) if qax_recipients else 0.0
    retry_total = sum(item.retry_count for item in recipients)

    task_trend = []
    for offset in range(6, -1, -1):
        day_start = today_start - timedelta(days=offset)
        day_end = day_start + timedelta(days=1)
        created_total = sum(1 for item in tasks if day_start <= item.created_at < day_end)
        completed_total = sum(
            1
            for event in status_events
            if event.to_status == "done" and day_start <= event.created_at < day_end
        )
        delayed_stock_total = sum(
            1
            for item in tasks
            if effective_task_delay_days(item, day_start) > 0
        )
        task_trend.append(
            {
                "label": day_start.strftime("%m-%d"),
                "created_total": created_total,
                "completed_total": completed_total,
                "delayed_total": delayed_stock_total,
            }
        )

    status_distribution = []
    for status in ("not_started", "in_progress", "done", "canceled"):
        status_distribution.append(
            {
                "key": status,
                "label": TASK_STATUS_LABELS.get(status, status),
                "value": sum(1 for item in tasks if item.main_status == status),
            }
        )

    priority_distribution = []
    for priority in ("high", "medium", "low"):
        priority_distribution.append(
            {
                "key": priority,
                "label": PRIORITY_LABELS.get(priority, priority),
                "value": sum(1 for item in tasks if item.priority == priority),
            }
        )

    owner_stats: dict[str, dict[str, int | str]] = {}
    warning_tasks = []
    for task in tasks:
        owner_member = next((member for member in task.members if member.member_role == "owner"), None)
        owner_name = owner_member.user.name if owner_member and owner_member.user else "未指定"
        stats = owner_stats.setdefault(
            owner_name,
            {
                "owner_name": owner_name,
                "task_total": 0,
                "done_total": 0,
                "delayed_total": 0,
                "due_soon_total": 0,
            },
        )
        stats["task_total"] = int(stats["task_total"]) + 1
        if task.main_status == "done":
            stats["done_total"] = int(stats["done_total"]) + 1
        if task.main_status not in OPEN_TASK_STATUSES:
            continue

        effective_delay_days = effective_task_delay_days(task, now)
        if effective_delay_days > 0:
            stats["delayed_total"] = int(stats["delayed_total"]) + 1
            warning_tasks.append(
                {
                    "task_id": task.id,
                    "title": task.title,
                    "owner_name": owner_name,
                    "end_at": task.end_at,
                    "delay_days": effective_delay_days,
                    "warning_type": "delayed",
                    "warning_text": f"已延期{effective_delay_days}天",
                    "route": f"/admin/tasks/{task.id}",
                }
            )
        elif is_task_due_soon(task, now):
            stats["due_soon_total"] = int(stats["due_soon_total"]) + 1
            warning_tasks.append(
                {
                    "task_id": task.id,
                    "title": task.title,
                    "owner_name": owner_name,
                    "end_at": task.end_at,
                    "delay_days": 0,
                    "warning_type": "due_soon",
                    "warning_text": "即将延期",
                    "route": f"/admin/tasks/{task.id}",
                }
            )

    owner_task_distribution = sorted(
        owner_stats.values(),
        key=lambda item: (-int(item["task_total"]), str(item["owner_name"])),
    )
    warning_tasks = sorted(
        warning_tasks,
        key=lambda item: (0 if item["warning_type"] == "delayed" else 1, item["end_at"]),
    )[:8]

    kpis = [
        {
            "label": "任务完成率",
            "value": completion_rate,
            "unit": "%",
            "detail": f"已完成 {done_total} / 全部 {len(tasks)}",
            "tone": "success" if completion_rate >= 70 else "warning" if completion_rate >= 40 else "danger",
        },
        {
            "label": "在途任务",
            "value": in_progress_total,
            "unit": "",
            "detail": f"未开始 {pending_total}，即将到期 {due_soon_total}",
            "tone": "primary",
        },
        {
            "label": "延期风险",
            "value": delayed_total,
            "unit": "",
            "detail": f"健康任务率 {healthy_task_rate}%",
            "tone": "danger" if delayed_total > 0 else "success",
        },
        {
            "label": "审批积压",
            "value": pending_delay_requests,
            "unit": "",
            "detail": f"沟通异常 {mail_failure_total}，重试 {retry_total}",
            "tone": "warning" if pending_delay_requests > 0 or mail_failure_total > 0 else "neutral",
        },
    ]

    attention_items = []
    if pending_delay_requests > 0:
        attention_items.append(
            {
                "title": "延期审批待处理",
                "description": "成员提交的延期申请还在等待管理员决策，优先清空积压可避免任务继续失真。",
                "value": f"{pending_delay_requests} 条",
                "tone": "warning",
                "route": "/admin/delay-requests",
                "action_label": "前往审批",
            }
        )
    if due_soon_total > 0:
        attention_items.append(
            {
                "title": "即将到期任务",
                "description": "未来 3 天内到期且仍未完结的任务需要尽快提醒或改派。",
                "value": f"{due_soon_total} 项",
                "tone": "danger",
                "route": "/admin/tasks",
                "action_label": "查看任务",
            }
        )
    if mail_failure_total > 0 or retry_total > 0:
        attention_items.append(
            {
                "title": "通知链路异常",
                "description": "存在发送失败、收件处理失败或重试累计，建议优先检查邮件与通知记录。",
                "value": f"{mail_failure_total} 异常 / {retry_total} 重试",
                "tone": "primary",
                "route": "/admin/notifications",
                "action_label": "检查通知",
            }
        )
    if delayed_total > 0:
        attention_items.append(
            {
                "title": "延期任务需要复盘",
                "description": "延期任务说明排期或协作存在阻塞，适合从高优先级任务先做排障。",
                "value": f"{delayed_total} 项",
                "tone": "danger",
                "route": "/admin/tasks",
                "action_label": "聚焦风险",
            }
        )
    if not attention_items:
        attention_items.append(
            {
                "title": "当前没有高优先级阻塞",
                "description": "审批、到期风险和通知异常都处在较平稳状态，可以继续推进计划型工作。",
                "value": "状态稳定",
                "tone": "success",
                "route": "/admin/tasks/new",
                "action_label": "新建任务",
            }
        )

    quick_actions = [
        {
            "title": "创建任务",
            "description": "直接发起新任务并同步通知成员。",
            "route": "/admin/tasks/new",
            "action_label": "立即创建",
            "tone": "primary",
        },
        {
            "title": "导入任务",
            "description": "批量导入计划，适合周计划或跨部门分发。",
            "route": "/admin/import-export",
            "action_label": "批量导入",
            "tone": "neutral",
        },
        {
            "title": "处理延期审批",
            "description": "优先消化待审批事项，避免任务时间继续漂移。",
            "route": "/admin/delay-requests",
            "action_label": "处理审批",
            "tone": "warning",
        },
        {
            "title": "检查通知链路",
            "description": "查看通知记录与发送反馈，确认成员是否真正收到提醒。",
            "route": "/admin/notifications",
            "action_label": "查看通知",
            "tone": "soft",
        },
    ]

    health_score = round(
        max(
            0,
            min(
                100,
                completion_rate * 0.35 + healthy_task_rate * 0.35 + email_success_rate * 0.15 + qax_delivery_rate * 0.15,
            ),
        )
    )

    return DashboardSummary(
        task_total=len(tasks),
        in_progress_total=in_progress_total,
        done_total=done_total,
        canceled_total=canceled_total,
        delayed_total=delayed_total,
        pending_total=pending_total,
        due_soon_total=due_soon_total,
        completion_rate=completion_rate,
        healthy_task_rate=healthy_task_rate,
        email_success_rate=email_success_rate,
        qax_delivery_rate=qax_delivery_rate,
        qax_read_rate=qax_read_rate,
        retry_total=retry_total,
        pending_delay_requests=pending_delay_requests,
        mail_failure_total=mail_failure_total,
        health_score=health_score,
        kpis=kpis,
        status_distribution=status_distribution,
        priority_distribution=priority_distribution,
        task_trend=task_trend,
        attention_items=attention_items,
        owner_task_distribution=owner_task_distribution,
        warning_tasks=warning_tasks,
        quick_actions=quick_actions,
    )


@router.post("/admin/users/{user_id}/disable", response_model=ApiMessage)
def disable_user(user_id: int, current_user: User = Depends(require_admin), db: Session = Depends(get_db)) -> ApiMessage:
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    ensure_last_admin_not_removed(db, user, user.role, False)
    user.is_active = False
    write_audit(
        db,
        current_user.id,
        "DISABLE_USER",
        "User",
        user.id,
        {"is_active": True},
        {"is_active": False},
        module_name="api.user",
        message=f"禁用用户「{user.name}」",
        detail={"username": user.username},
    )
    db.commit()
    return ApiMessage(message="用户已禁用")


@router.post("/admin/users/{user_id}/enable", response_model=ApiMessage)
def enable_user(user_id: int, current_user: User = Depends(require_admin), db: Session = Depends(get_db)) -> ApiMessage:
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    user.is_active = True
    write_audit(
        db,
        current_user.id,
        "ENABLE_USER",
        "User",
        user.id,
        {"is_active": False},
        {"is_active": True},
        module_name="api.user",
        message=f"启用用户「{user.name}」",
        detail={"username": user.username},
    )
    db.commit()
    return ApiMessage(message="用户已启用")


@router.post("/admin/users/{user_id}/reset-password", response_model=ApiMessage)
def reset_password(user_id: int, current_user: User = Depends(require_admin), db: Session = Depends(get_db)) -> ApiMessage:
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    user.password_hash = build_default_password_hash(settings.default_password)
    write_audit(
        db,
        current_user.id,
        "RESET_PASSWORD",
        "User",
        user.id,
        {},
        {"default_password_applied": True},
        module_name="api.user",
        message=f"重置用户「{user.name}」密码",
        detail={"username": user.username},
    )
    db.commit()
    return ApiMessage(message="密码已重置")


@router.get("/tasks", response_model=list[TaskOut])
def list_tasks(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> list[TaskOut]:
    """获取任务列表。

    管理员可查看全部未删除任务，成员仅可查看自己参与的任务。
    """
    query = db.query(Task).filter(Task.deleted_at.is_(None))
    if not is_admin_role(current_user.role):
        task_ids = [tm.task_id for tm in db.query(TaskMember).filter(TaskMember.user_id == current_user.id).all()]
        query = query.filter(Task.id.in_(task_ids))
    return [serialize_task(task, db) for task in query.order_by(Task.id.desc()).all()]


@router.post("/tasks", response_model=TaskOut)
def create_task(payload: TaskCreate, current_user: User = Depends(require_admin), db: Session = Depends(get_db)) -> TaskOut:
    """创建任务并立即返回，任务创建通知交由后台发送。"""
    task = _create_task_record(payload, current_user, db, source="web")
    db.commit()
    db.refresh(task)
    task_out = serialize_task(task, db)
    _enqueue_task_created_notifications(task.id)
    return task_out


@router.get("/tasks/import-template")
def task_import_template(_: User = Depends(require_admin)) -> StreamingResponse:
    """生成任务导入 Excel 模板。

    模板会同时提供：
    - `任务导入模板` 工作表：用于直接录入导入数据；
    - `填写说明` 工作表：说明字段含义、格式与示例规则。
    """
    config_template = task_import_template_file()
    if config_template is not None:
        return FileResponse(
            config_template,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="task-import-template.xlsx",
        )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "任务导入模板"
    sheet.append(["序号", "任务名称", "任务内容", "子任务", "负责人", "截止日期"])
    sheet.append([1, "地市数据分域", "把数据分域工作落到实处，明确各市的数据管理范围与权限。", "制定地市数据分域清单与规则", "张三、李四", datetime(2026, 12, 31)])
    sheet.append([1, "地市数据分域", "把数据分域工作落到实处，明确各市的数据管理范围与权限。", "部署地市数据分域访问控制", "李四", datetime(2026, 12, 31)])
    sheet.append([2, "数据治理", "依据数据敏感程度进行分级，并制定相应的安全策略与访问权限控制，确保客户隐私数据与核心商业数据在存储、流转和使用中的安全合规。", "", "张三、李四、老六", datetime(2026, 5, 31)])
    sheet.append([3, "……", "", "", "", ""])
    sheet.append(["", "", "", "", "", ""])
    sheet.append(["", "", "", "", "", ""])
    sheet.append([
        "补充说明：\n1.同一个任务如果有不同的子任务，每个子任务单独占一行，但序号相同。\n2.每行中的负责人如果有多个，以中文顿号或逗号区分，通知需要按人名进行发送。\n3.如果同一个任务下，同一个人涉及多个子任务，则将子任务合并为一个通知。\n4.邮件内容需简化，任务标题和子任务按截止日期聚合展示。\n5.桌面消息的标题和内容与邮件一致",
        "",
        "",
        "",
        "",
        "",
    ])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_index in range(2, 9):
        for column_index in range(1, 7):
            sheet.cell(row=row_index, column=column_index).alignment = Alignment(vertical="top", wrap_text=True)
    for column, width in {"A": 10, "B": 24, "C": 48, "D": 30, "E": 24, "F": 18}.items():
        sheet.column_dimensions[column].width = width
    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    return StreamingResponse(
        iter([excel_buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="task-import-template.xlsx"'},
    )
    sheet = workbook.active
    sheet.title = "任务导入模板"

    headers = [
        "任务标题(title)",
        "任务内容(content)",
        "负责人姓名(owner_name)",
        "参与人员姓名(participant_names)",
        "开始时间(start_at)",
        "结束时间(end_at)",
        "优先级(priority)",
        "备注(remark)",
        "到期前提醒天数(due_remind_days)",
        "里程碑名称(milestone_names)",
        "里程碑时间(milestone_datetimes)",
        "里程碑提醒天数(remind_offsets)",
        "子任务标题(subtask_titles)",
        "子任务内容(subtask_contents)",
        "子任务执行人姓名(subtask_assignee_names)",
    ]
    example_row = [
        "Q2 重点项目汇总",
        "完成季度重点项目进展汇总与风险登记。",
        "系统管理员",
        "默认成员,测试成员",
        "2026-04-23T09:00:00",
        "2026-04-30T18:00:00",
        "high",
        "用于周会汇报",
        2,
        "初稿整理|主管复核|正式提交",
        "2026-04-24T18:00:00|2026-04-27T18:00:00|2026-04-30T12:00:00",
        "1,2|1|1,3",
        "汇总原始数据|复核内容|提交汇报",
        "整理各部门数据|检查格式与风险项|输出最终稿",
        "默认成员|测试成员|系统管理员",
    ]

    sheet.append(headers)
    sheet.append(example_row)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    column_widths = {
        "A": 22,
        "B": 36,
        "C": 20,
        "D": 28,
        "E": 22,
        "F": 22,
        "G": 14,
        "H": 24,
        "I": 18,
        "J": 30,
        "K": 42,
        "L": 28,
        "M": 30,
        "N": 42,
        "O": 30,
    }
    for column, width in column_widths.items():
        sheet.column_dimensions[column].width = width

    instruction_sheet = workbook.create_sheet("填写说明")
    instruction_sheet.append(["字段", "是否必填", "填写说明", "示例"])
    instructions = [
        ["任务标题(title)", "是", "任务名称，建议直接对应业务事项标题。", "Q2 重点项目汇总"],
        ["任务内容(content)", "是", "任务详细说明。", "完成季度重点项目进展汇总与风险登记。"],
        ["负责人姓名(owner_name)", "是", "填写系统中的姓名，不能重名。", "系统管理员"],
        ["参与人员姓名(participant_names)", "否", "多个姓名用英文逗号分隔，不能重名。", "默认成员,测试成员"],
        ["开始时间(start_at)", "是", "使用 ISO 格式时间，精确到秒。", "2026-04-23T09:00:00"],
        ["结束时间(end_at)", "是", "使用 ISO 格式时间，必须晚于开始时间。", "2026-04-30T18:00:00"],
        ["优先级(priority)", "是", "仅支持 high、medium、low。", "high"],
        ["备注(remark)", "否", "补充说明，可留空。", "用于周会汇报"],
        ["到期前提醒天数(due_remind_days)", "否", "填写整数，0 表示不自动提醒。", "2"],
        ["里程碑名称(milestone_names)", "否", "多个里程碑用 | 分隔，顺序需与时间、提醒列一致。", "初稿整理|主管复核|正式提交"],
        ["里程碑时间(milestone_datetimes)", "否", "多个时间用 | 分隔，格式需与任务时间一致。", "2026-04-24T18:00:00|2026-04-27T18:00:00|2026-04-30T12:00:00"],
        ["里程碑提醒天数(remind_offsets)", "否", "每个里程碑的提醒天数用英文逗号分隔，多个里程碑之间用 | 分隔。", "1,2|1|1,3"],
        ["子任务标题(subtask_titles)", "否", "多个子任务标题用 | 分隔。", "汇总原始数据|复核内容|提交汇报"],
        ["子任务内容(subtask_contents)", "否", "多个子任务内容用 | 分隔，顺序需与子任务标题一致。", "整理各部门数据|检查格式与风险项|输出最终稿"],
        ["子任务执行人姓名(subtask_assignee_names)", "否", "填写系统姓名，多个执行人用 | 分隔，且必须属于主任务成员，不能重名。", "默认成员|测试成员|系统管理员"],
    ]
    for row in instructions:
        instruction_sheet.append(row)

    for cell in instruction_sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for column in ("A", "B", "C", "D"):
        instruction_sheet.column_dimensions[column].width = 32 if column == "C" else 24

    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    response_headers = {
        "Content-Disposition": 'attachment; filename="task-import-template.xlsx"',
    }
    return StreamingResponse(
        iter([excel_buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=response_headers,
    )


@router.get("/tasks/import-histories", response_model=list[TaskImportHistoryOut])
def list_task_import_histories(_: User = Depends(require_admin), db: Session = Depends(get_db)) -> list[TaskImportHistoryOut]:
    """返回最近的任务导入历史。"""
    histories = db.query(TaskImportHistory).order_by(TaskImportHistory.id.desc()).limit(20).all()
    return [serialize_task_import_history(item, db) for item in histories]


@router.get("/tasks/{task_id}/notification-preview", response_model=NotificationPreviewOut)
def get_task_notification_preview(
    task_id: int,
    channel: str = Query("email"),
    notify_type: str = Query("task_created"),
    recipient_user_id: int | None = Query(None),
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> NotificationPreviewOut:
    """棰勮鏌愪釜浠诲姟鍦ㄦ寚瀹氭笭閬撱€佹寚瀹氭帴鏀朵汉涓嬫渶缁堜細鍙戦€佺殑閫氱煡鍐呭銆?"""
    if channel not in {"email", "qax"}:
        raise HTTPException(status_code=400, detail="浠呮敮鎸侀偖浠朵笌鍗虫椂娑堟伅棰勮")
    if notify_type not in {"task_created", "manual_remind", "due_remind"}:
        raise HTTPException(status_code=400, detail="褰撳墠棰勮浠呮敮鎸佷换鍔″垱寤恒€佹墜鍔ㄦ彁閱掑拰鍒版湡鎻愰啋")
    task = db.query(Task).filter(Task.id == task_id, Task.deleted_at.is_(None)).first()
    if not task:
        raise HTTPException(status_code=404, detail="浠诲姟涓嶅瓨鍦?")
    recipient = None
    if recipient_user_id is not None:
        membership = db.query(TaskMember).filter(TaskMember.task_id == task.id, TaskMember.user_id == recipient_user_id).first()
        if not membership:
            raise HTTPException(status_code=400, detail="棰勮鎺ユ敹浜哄繀椤诲睘浜庡綋鍓嶄换鍔℃垚鍛?")
        recipient = membership.user if membership.user else db.query(User).filter(User.id == recipient_user_id).first()
    preview = preview_notification_content(
        db=db,
        task=task,
        channel=channel,
        notify_type=notify_type,
        recipient=recipient,
    )
    return NotificationPreviewOut(
        channel=channel,
        channel_text=notification_channel_text(channel),
        notify_type=notify_type,
        notify_type_text=NOTIFICATION_TYPE_LABELS.get(notify_type, notify_type),
        recipient_user_id=recipient.id if recipient else None,
        recipient_name=recipient.name if recipient else "",
        recipient_email=recipient.email if recipient else "",
        template_name=str(preview.get("template_name") or ""),
        template_version=preview.get("template_version"),
        subject=str(preview.get("subject") or ""),
        content=str(preview.get("content") or ""),
        context=preview.get("context") or {},
    )


@router.get("/tasks/{task_id}", response_model=TaskDetailOut)
def get_task(task_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> TaskDetailOut:
    task = db.query(Task).filter(Task.id == task_id, Task.deleted_at.is_(None)).first()
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    if not is_admin_role(current_user.role):
        membership = db.query(TaskMember).filter(TaskMember.task_id == task_id, TaskMember.user_id == current_user.id).first()
        if not membership:
            raise HTTPException(status_code=403, detail="无权访问该任务")
    members = db.query(TaskMember).filter(TaskMember.task_id == task.id).all()
    milestones = db.query(TaskMilestone).filter(TaskMilestone.task_id == task.id).order_by(TaskMilestone.sort_order.asc()).all()
    subtasks = db.query(TaskSubtask).filter(TaskSubtask.task_id == task.id).order_by(TaskSubtask.sort_order.asc()).all()
    notifications = db.query(Notification).filter(Notification.task_id == task.id).order_by(Notification.id.desc()).all()
    delay_requests = db.query(DelayRequest).filter(DelayRequest.task_id == task.id).order_by(DelayRequest.id.desc()).all()
    events = db.query(TaskStatusEvent).filter(TaskStatusEvent.task_id == task.id).order_by(TaskStatusEvent.id.desc()).all()
    base = serialize_task(task, db)
    return TaskDetailOut(
        **base.model_dump(),
        members=[
            {
                "id": item.id,
                "user_id": item.user_id,
                "name": item.user.name if item.user else "",
                "email": item.user.email if item.user else "",
                "member_role": item.member_role,
                "member_role_text": MEMBER_ROLE_LABELS.get(item.member_role, item.member_role),
                "latest_notifications": _task_latest_notifications(db, task.id, user_id=item.user_id),
                "display_role_text": "负责人",
            }
            for item in members
        ],
        milestones=[
            {
                "id": item.id,
                "name": item.name,
                "planned_at": item.planned_at,
                "remind_offsets": [int(value) for value in item.remind_offsets.split(",") if value],
                "status": item.status,
            }
            for item in milestones
        ],
        subtasks=[
            {
                "id": item.id,
                "title": item.title,
                "content": item.content,
                "assignee_id": item.assignee_id,
                "assignee_name": item.assignee.name if item.assignee else "",
                "assignee_email": item.assignee.email if item.assignee else "",
                "status": item.status,
                "status_text": SUBTASK_STATUS_LABELS.get(item.status, item.status),
                "sort_order": item.sort_order,
                "latest_notifications": _task_latest_notifications(db, task.id, user_id=item.assignee_id),
            }
            for item in subtasks
        ],
        notifications=[serialize_notification(item, db).model_dump() for item in notifications],
        delay_requests=[
            {
                "id": item.id,
                "applicant_id": item.applicant_id,
                "applicant_name": db.query(User).filter(User.id == item.applicant_id).first().name if item.applicant_id else "",
                "approver_id": item.approver_id,
                "approver_name": db.query(User).filter(User.id == item.approver_id).first().name if item.approver_id else "",
                "apply_reason": item.apply_reason,
                "approval_status": item.approval_status,
                "approval_status_text": DELAY_STATUS_LABELS.get(item.approval_status, item.approval_status),
                "original_deadline": item.original_deadline,
                "proposed_deadline": item.proposed_deadline,
                "approved_deadline": item.approved_deadline,
                "approve_remark": item.approve_remark,
                "decided_by_channel": item.decided_by_channel,
                "decided_at": item.decided_at,
                "version": item.version,
            }
            for item in delay_requests
        ],
        events=[
            {
                "id": item.id,
                "from_status": item.from_status,
                "to_status": item.to_status,
                "from_status_text": task_status_text(item.from_status),
                "to_status_text": task_status_text(item.to_status),
                "source": item.source,
                "remark": item.remark,
                "created_at": item.created_at,
            }
            for item in events
        ],
    )


@router.put("/tasks/{task_id}", response_model=TaskOut)
def update_task(task_id: int, payload: TaskCreate, current_user: User = Depends(require_admin), db: Session = Depends(get_db)) -> TaskOut:
    task = db.query(Task).filter(Task.id == task_id, Task.deleted_at.is_(None)).first()
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    if payload.start_at > payload.end_at:
        raise HTTPException(status_code=400, detail="开始时间不能晚于结束时间")

    member_ids = {payload.owner_id, *payload.participant_ids}
    for subtask in payload.subtasks:
        if subtask.assignee_id not in member_ids:
            raise HTTPException(status_code=400, detail="子任务执行人必须从主任务参与成员中选择")

    before = {"title": task.title, "end_at": task.end_at.isoformat(), "priority": task.priority}
    task.title = payload.title
    task.content = payload.content
    task.priority = payload.priority
    task.remark = payload.remark
    task.start_at = payload.start_at
    task.end_at = payload.end_at
    task.due_remind_days = max(payload.due_remind_days, 0)
    task.planned_minutes = int((payload.end_at - payload.start_at).total_seconds() // 60)

    db.query(TaskMember).filter(TaskMember.task_id == task.id).delete()
    for user_id in member_ids:
        member_role = "owner" if user_id == payload.owner_id else "participant"
        db.add(TaskMember(task_id=task.id, user_id=user_id, member_role=member_role))

    db.query(TaskMilestone).filter(TaskMilestone.task_id == task.id).delete()
    for milestone in payload.milestones:
        if milestone.planned_at < payload.start_at or milestone.planned_at > payload.end_at:
            raise HTTPException(status_code=400, detail="里程碑时间必须位于任务时间范围内")
        db.add(
            TaskMilestone(
                task_id=task.id,
                name=milestone.name,
                planned_at=milestone.planned_at,
                remind_offsets=",".join(str(item) for item in milestone.remind_offsets),
                sort_order=milestone.sort_order,
            )
        )

    db.query(TaskSubtask).filter(TaskSubtask.task_id == task.id).delete()
    for subtask in payload.subtasks:
        db.add(
            TaskSubtask(
                task_id=task.id,
                title=subtask.title,
                content=subtask.content,
                assignee_id=subtask.assignee_id,
                sort_order=subtask.sort_order,
                status=subtask.status or "pending",
            )
        )
    write_audit(
        db,
        current_user.id,
        "UPDATE_TASK",
        "Task",
        task.id,
        before,
        {"title": task.title, "end_at": task.end_at.isoformat()},
        module_name="api.task",
        message=f"更新任务《{task.title}》",
        detail={
            "start_at": task.start_at.isoformat(),
            "end_at": task.end_at.isoformat(),
            "owner_id": next((item.user_id for item in task.members if item.member_role == "owner"), None),
            "participant_total": sum(1 for item in task.members if item.member_role != "owner"),
        },
    )
    db.commit()
    db.refresh(task)
    return serialize_task(task, db)


@router.delete("/tasks/{task_id}", response_model=ApiMessage)
def delete_task(task_id: int, current_user: User = Depends(require_admin), db: Session = Depends(get_db)) -> ApiMessage:
    task = db.query(Task).filter(Task.id == task_id, Task.deleted_at.is_(None)).first()
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    cleaned_reminders = cleanup_task_scheduled_notifications(db, task.id)
    task.due_remind_days = 0
    task.deleted_at = shanghai_now_naive()
    write_audit(
        db,
        current_user.id,
        "DELETE_TASK",
        "Task",
        task.id,
        {"deleted_at": None},
        {"deleted_at": task.deleted_at.isoformat(), "cleaned_due_reminders": cleaned_reminders},
        module_name="api.task",
        message=f"删除任务《{task.title}》",
        detail={"cleaned_due_reminders": cleaned_reminders, "due_remind_days": task.due_remind_days},
    )
    db.commit()
    return ApiMessage(message="任务已删除")


@router.post("/tasks/{task_id}/status", response_model=TaskOut)
def change_status(task_id: int, payload: TaskStatusUpdate, current_user: User = Depends(require_admin), db: Session = Depends(get_db)) -> TaskOut:
    task = db.query(Task).filter(Task.id == task_id, Task.deleted_at.is_(None)).first()
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    before = {"main_status": task.main_status}
    task.main_status = payload.main_status
    now = shanghai_now_naive()
    if payload.main_status == "done":
        if task.actual_minutes == 0:
            task.actual_minutes = int((now - task.start_at).total_seconds() // 60)
        if task.completed_at is None:
            task.completed_at = now
    elif before["main_status"] == "done":
        task.completed_at = None
    db.add(TaskStatusEvent(task_id=task.id, from_status=before["main_status"], to_status=task.main_status, source="web", remark=payload.remark, operator_id=current_user.id))
    write_audit(
        db,
        current_user.id,
        "CHANGE_TASK_STATUS",
        "Task",
        task.id,
        before,
        {"main_status": task.main_status},
        module_name="api.task",
        message=f"更新任务《{task.title}》状态为 {task.main_status}",
        detail={"remark": payload.remark},
    )
    db.commit()
    db.refresh(task)
    return serialize_task(task, db)


@router.post("/tasks/{task_id}/lock", response_model=ApiMessage)
def lock_task(task_id: int, current_user: User = Depends(require_admin), db: Session = Depends(get_db)) -> ApiMessage:
    task = db.query(Task).filter(Task.id == task_id, Task.deleted_at.is_(None)).first()
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    task.state_locked = True
    db.commit()
    return ApiMessage(message="任务状态已锁定")


@router.post("/tasks/{task_id}/unlock", response_model=ApiMessage)
def unlock_task(task_id: int, current_user: User = Depends(require_admin), db: Session = Depends(get_db)) -> ApiMessage:
    task = db.query(Task).filter(Task.id == task_id, Task.deleted_at.is_(None)).first()
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    task.state_locked = False
    db.commit()
    return ApiMessage(message="任务状态已解锁")


def _send_task_reminders(
    db: Session,
    task: Task,
    notify_type: str,
    extra_context: dict[str, str] | None = None,
    recipient_user_ids: list[int] | None = None,
    channels: list[str] | None = None,
) -> None:
    """统一发送任务相关提醒，保证邮件与即时消息正文保持同一份上下文。"""
    target_channels = channels or ["email", "qax"]
    for channel in target_channels:
        create_notification_with_recipients(
            db,
            task.id,
            channel,
            notify_type,
            "",
            recipient_user_ids=recipient_user_ids,
            extra_context=extra_context,
        )


@router.get("/templates", response_model=list[dict])
def list_templates(template_kind: str | None = Query(None), _: User = Depends(require_admin), db: Session = Depends(get_db)) -> list[dict]:
    query = db.query(Template)
    if template_kind:
        query = query.filter(Template.template_kind == template_kind)
    return [
        {
            "id": item.id,
            "name": item.name,
            "template_kind": item.template_kind,
            "notify_type": item.notify_type,
            "notify_type_text": NOTIFICATION_TYPE_LABELS.get(item.notify_type, item.notify_type),
            "priority": item.priority,
            "version": item.version,
            "enabled": item.enabled,
            "is_default": item.is_default,
            "subject_rule": item.subject_rule,
            "body_rule": item.body_rule,
            "content": item.content,
        }
        for item in query.order_by(Template.id.desc()).all()
    ]


@router.get("/templates/options", response_model=dict)
def template_options(_: User = Depends(require_admin)) -> dict:
    return {
        "template_kind_options": [
            {"value": "MAIL_SEND", "label": "邮件发送模板"},
            {"value": "QAX_SEND", "label": "即时消息发送模板"},
            {"value": "MAIL_REPLY", "label": "邮件回复模板"},
        ],
        "notify_type_options": {
            kind: [{"value": code, "label": NOTIFICATION_TYPE_LABELS.get(code, code)} for code in codes]
            for kind, codes in TEMPLATE_NOTIFY_TYPE_OPTIONS.items()
        },
    }


def _validate_template_notify_type(template_kind: str, notify_type: str) -> None:
    allowed = TEMPLATE_NOTIFY_TYPE_OPTIONS.get(template_kind, [])
    if notify_type not in allowed:
        raise HTTPException(status_code=400, detail="通知类型与模板类型不匹配")


@router.post("/templates", response_model=dict)
def create_template(payload: TemplateCreate, current_user: User = Depends(require_admin), db: Session = Depends(get_db)) -> dict:
    _validate_template_notify_type(payload.template_kind, payload.notify_type)
    validate_template_content(payload.template_kind, payload.notify_type, payload.content)
    template = Template(**payload.model_dump())
    db.add(template)
    db.flush()
    write_audit(
        db,
        current_user.id,
        "CREATE_TEMPLATE",
        "Template",
        template.id,
        {},
        {"name": template.name},
        module_name="api.template",
        message=f"创建模板《{template.name}》",
        detail={"template_kind": template.template_kind, "notify_type": template.notify_type, "version": template.version},
    )
    db.commit()
    return {"id": template.id}


@router.put("/templates/{template_id}", response_model=dict)
def update_template(template_id: int, payload: TemplateCreate, current_user: User = Depends(require_admin), db: Session = Depends(get_db)) -> dict:
    template = db.query(Template).filter(Template.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在")
    _validate_template_notify_type(payload.template_kind, payload.notify_type)
    validate_template_content(payload.template_kind, payload.notify_type, payload.content)
    before = {"name": template.name, "version": template.version}
    for key, value in payload.model_dump().items():
        setattr(template, key, value)
    write_audit(
        db,
        current_user.id,
        "UPDATE_TEMPLATE",
        "Template",
        template.id,
        before,
        {"name": template.name, "version": template.version},
        module_name="api.template",
        message=f"更新模板《{template.name}》",
        detail={"template_kind": template.template_kind, "notify_type": template.notify_type, "enabled": template.enabled},
    )
    db.commit()
    return {"id": template.id}


@router.post("/templates/{template_id}/set-default", response_model=ApiMessage)
def set_default_template(template_id: int, current_user: User = Depends(require_admin), db: Session = Depends(get_db)) -> ApiMessage:
    template = db.query(Template).filter(Template.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在")
    db.query(Template).filter(
        Template.template_kind == template.template_kind,
        Template.notify_type == template.notify_type,
        Template.is_default.is_(True),
    ).update({"is_default": False})
    template.is_default = True
    write_audit(
        db,
        current_user.id,
        "SET_TEMPLATE_DEFAULT",
        "Template",
        template.id,
        {},
        {"is_default": True},
        module_name="api.template",
        message=f"设置模板《{template.name}》为默认模板",
        detail={"template_kind": template.template_kind, "notify_type": template.notify_type},
    )
    db.commit()
    return ApiMessage(message="模板默认项已更新")


@router.post("/templates/preview-match", response_model=dict)
def preview_match(payload: TemplatePreviewRequest, _: User = Depends(require_admin), db: Session = Depends(get_db)) -> dict:
    candidates = [
        item for item in sort_templates(db.query(Template).filter(Template.template_kind == payload.template_kind, Template.enabled.is_(True)).all())
        if template_matches(item, payload.subject, payload.body)
    ]
    return {
        "matches": [{"id": item.id, "name": item.name, "version": item.version, "priority": item.priority} for item in candidates],
        "selected": candidates[0].id if candidates else None,
    }


@router.get("/notifications", response_model=list[NotificationOut])
def list_notifications(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> list[NotificationOut]:
    query = db.query(Notification)
    if not is_admin_role(current_user.role):
        task_ids = [tm.task_id for tm in db.query(TaskMember).filter(TaskMember.user_id == current_user.id).all()]
        query = query.filter(Notification.task_id.in_(task_ids))
    return [serialize_notification(item, db) for item in query.order_by(Notification.id.desc()).all()]


@router.get("/notifications/{notification_id}", response_model=NotificationDetailOut)
def get_notification_detail(notification_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> NotificationDetailOut:
    """获取通知详情，包括正文和按成员拆分的送达情况。"""
    notification = db.query(Notification).filter(Notification.id == notification_id).first()
    if not notification:
        raise HTTPException(status_code=404, detail="通知不存在")
    ensure_notification_access(notification, current_user, db)
    return serialize_notification_detail(notification, db)


@router.get("/admin/mail/events", response_model=list[MailEventOut])
def list_mail_events(_: User = Depends(require_admin), db: Session = Depends(get_db)) -> list[MailEventOut]:
    """获取已匹配模板的邮件列表。"""
    query = (
        db.query(MailEvent)
        .filter(MailEvent.resolved_template_id.isnot(None))
        .order_by(MailEvent.id.desc())
        .limit(100)
    )
    return [serialize_mail_event(item, db) for item in query.all()]


@router.get("/admin/mail/events/{event_id}", response_model=MailEventDetailOut)
def get_mail_event_detail(event_id: int, _: User = Depends(require_admin), db: Session = Depends(get_db)) -> MailEventDetailOut:
    """获取邮件详情页所需的完整记录。"""
    event = db.query(MailEvent).filter(MailEvent.id == event_id, MailEvent.resolved_template_id.isnot(None)).first()
    if not event:
        raise HTTPException(status_code=404, detail="邮件记录不存在")
    return serialize_mail_event_detail(event, db)


@router.get("/admin/mail/poll-state", response_model=MailPollStateOut)
def get_mail_poll_state(_: User = Depends(require_admin), db: Session = Depends(get_db)) -> MailPollStateOut:
    """返回自动收件状态与下一次预计执行时间。"""
    return serialize_mail_poll_state(db)


@router.get("/admin/scheduler/settings", response_model=dict)
def get_scheduler_settings(_: User = Depends(require_admin)) -> dict:
    return runtime_settings_dict()


@router.put("/admin/scheduler/settings", response_model=dict)
def update_scheduler_settings(payload: RuntimeSettingsUpdate, current_user: User = Depends(require_admin), db: Session = Depends(get_db)) -> dict:
    saved = save_runtime_settings(
        {
            "mail_auto_poll_enabled": payload.mail_auto_poll_enabled,
            "mail_auto_poll_interval_seconds": payload.mail_auto_poll_interval_seconds,
            "mail_inbox_max_scan": payload.mail_inbox_max_scan,
            "due_remind_enabled": payload.due_remind_enabled,
            "due_remind_run_at": payload.due_remind_run_at,
            "overdue_remind_enabled": payload.overdue_remind_enabled,
            "overdue_remind_run_at": payload.overdue_remind_run_at,
            "qax_auto_collect_enabled": payload.qax_auto_collect_enabled,
            "qax_auto_collect_interval_seconds": payload.qax_auto_collect_interval_seconds,
            "mail_scan_baseline_at": payload.mail_scan_baseline_at.isoformat() if payload.mail_scan_baseline_at else "",
            "qax_browser_visible": payload.qax_browser_visible,
            "qax_base_url": payload.qax_base_url,
            "qax_username": payload.qax_username,
            "qax_password": payload.qax_password,
            "qax_group_name": payload.qax_group_name,
            "qax_ignore_https_errors": payload.qax_ignore_https_errors,
            "smtp_host": payload.smtp_host,
            "smtp_port": payload.smtp_port,
            "smtp_user": payload.smtp_user,
            "smtp_password": payload.smtp_password,
            "smtp_from_address": payload.smtp_from_address,
            "smtp_use_tls": payload.smtp_use_tls,
            "smtp_use_ssl": payload.smtp_use_ssl,
            "smtp_timeout_seconds": payload.smtp_timeout_seconds,
            "mail_inbox_protocol": payload.mail_inbox_protocol,
            "imap_host": payload.imap_host,
            "imap_port": payload.imap_port,
            "imap_user": payload.imap_user,
            "imap_password": payload.imap_password,
            "imap_use_tls": payload.imap_use_tls,
            "imap_use_ssl": payload.imap_use_ssl,
            "pop3_host": payload.pop3_host,
            "pop3_port": payload.pop3_port,
            "pop3_user": payload.pop3_user,
            "pop3_password": payload.pop3_password,
            "pop3_use_tls": payload.pop3_use_tls,
            "pop3_use_ssl": payload.pop3_use_ssl,
        }
    )
    data = runtime_settings_dict()
    write_audit(
        db,
        current_user.id,
        "UPDATE_SCHEDULER_SETTINGS",
        "RuntimeSettings",
        None,
        {},
        data,
        module_name="api.scheduler",
        message="更新定时任务设置",
        detail=data,
    )
    db.commit()
    return data


@router.post("/admin/mail/test", response_model=dict)
def test_mail_settings(_: User = Depends(require_admin)) -> dict:
    return diagnose_mail_settings()


@router.post("/admin/mail/inbox-test", response_model=dict)
def test_mail_inbox(_: User = Depends(require_admin)) -> dict:
    """测试当前启用的收件协议配置。"""
    return diagnose_inbox_settings()


@router.post("/admin/mail/poll", response_model=dict)
def poll_mail_inbox(current_user: User = Depends(require_admin), db: Session = Depends(get_db)) -> dict:
    result = start_collect("mail")
    write_audit(
        db,
        current_user.id,
        "MANUAL_MAIL_POLL",
        "MailInbox",
        None,
        {},
        {"status": result.get("status"), "accepted": result.get("accepted", False)},
        log_level="INFO" if result.get("accepted") else "WARNING",
        module_name="api.mail",
        message="手动邮件采集已转入后台执行" if result.get("accepted") else "已有后台采集正在执行，忽略本次邮件采集",
        detail=result,
    )
    db.commit()
    return result


@router.post("/admin/mail/baseline", response_model=dict)
def reset_mail_baseline(payload: MailBaselineRequest | None = None, current_user: User = Depends(require_admin), db: Session = Depends(get_db)) -> dict:
    baseline_at = payload.baseline_at if payload else None
    result = initialize_mail_scan_baseline(db, baseline_at)
    write_audit(
        db,
        current_user.id,
        "RESET_MAIL_BASELINE",
        "MailScanState",
        1,
        {},
        {"baseline_at": baseline_at.isoformat() if baseline_at else ""},
        module_name="api.mail",
        message="设置邮件扫描基准时间",
        detail=result,
    )
    db.commit()
    return result


@router.post("/admin/qax/collect", response_model=dict)
def run_qax_collect(current_user: User = Depends(require_admin), db: Session = Depends(get_db)) -> dict:
    """手动触发一次 QAX 状态采集。"""
    result = start_collect("qax")
    write_audit(
        db,
        current_user.id,
        "MANUAL_COLLECT_QAX_STATUS",
        "Notification",
        None,
        {},
        {"status": result.get("status"), "accepted": result.get("accepted", False)},
        log_level="INFO" if result.get("accepted") else "WARNING",
        module_name="api.qax",
        message="手动 QAX 采集已转入后台执行" if result.get("accepted") else "已有后台采集正在执行，忽略本次 QAX 采集",
        detail=result,
    )
    db.commit()
    return result


@router.get("/admin/collect/state", response_model=dict)
def get_collect_state(task_id: int | None = Query(default=None), _: User = Depends(require_admin)) -> dict:
    """返回邮件与 QAX 后台采集状态，供前端展示等待进度。"""

    return collect_state(task_id)


@router.post("/tasks/{task_id}/sync-collect", response_model=dict)
def sync_collect_task(task_id: int, current_user: User = Depends(require_admin), db: Session = Depends(get_db)) -> dict:
    """按任务触发邮件与 QAX 后台同步采集，并按参与人展示进度。"""

    task = db.query(Task).filter(Task.id == task_id, Task.deleted_at.is_(None)).first()
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    result = start_collect("sync", task_id)
    write_audit(
        db,
        current_user.id,
        "SYNC_COLLECT_TASK",
        "Task",
        task.id,
        {},
        {"status": result.get("status"), "accepted": result.get("accepted", False)},
        log_level="INFO" if result.get("accepted") else "WARNING",
        module_name="api.collect",
        message="任务同步采集已转入后台执行" if result.get("accepted") else "已有后台采集正在执行，忽略本次任务同步",
        detail=result,
    )
    db.commit()
    return result


@router.post("/tasks/{task_id}/remind", response_model=ApiMessage)
def remind_task(task_id: int, current_user: User = Depends(require_admin), db: Session = Depends(get_db)) -> ApiMessage:
    task = db.query(Task).filter(Task.id == task_id, Task.deleted_at.is_(None)).first()
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    _send_task_reminders(
        db,
        task,
        "manual_remind",
        extra_context={"remind_focus": "主任务整体进度跟进"},
    )
    write_audit(
        db,
        current_user.id,
        "REMIND_TASK",
        "Task",
        task.id,
        {},
        {"notification_created": True, "remind_focus": "主任务整体进度跟进"},
    )
    db.commit()
    return ApiMessage(message="已创建提醒通知")


@router.post("/tasks/{task_id}/subtasks/{subtask_id}/remind", response_model=ApiMessage)
def remind_task_subtask(
    task_id: int,
    subtask_id: int,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiMessage:
    """按子任务定向发送提醒，仅触达当前子任务执行人。"""
    task = db.query(Task).filter(Task.id == task_id, Task.deleted_at.is_(None)).first()
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    subtask = (
        db.query(TaskSubtask)
        .filter(TaskSubtask.id == subtask_id, TaskSubtask.task_id == task.id)
        .first()
    )
    if not subtask:
        raise HTTPException(status_code=404, detail="子任务不存在")
    if subtask.status == "done":
        raise HTTPException(status_code=400, detail="已完成的子任务无需再次提醒")

    assignee_name = subtask.assignee.name if subtask.assignee else f"成员#{subtask.assignee_id}"
    focus_text = f"请优先跟进子任务“{subtask.title}”"
    subtask_summary = f"1. {subtask.title}（执行人：{assignee_name}）"
    if subtask.content:
        subtask_summary = f"{subtask_summary}：{subtask.content}"
    _send_task_reminders(
        db,
        task,
        "manual_remind",
        recipient_user_ids=[subtask.assignee_id],
        extra_context={
            "remind_focus": focus_text,
            "subtask_summary": subtask_summary,
            "subtask_brief": subtask_summary,
        },
    )
    write_audit(
        db,
        current_user.id,
        "REMIND_SUBTASK",
        "TaskSubtask",
        subtask.id,
        {},
        {"task_id": task.id, "notification_created": True, "remind_focus": focus_text},
    )
    db.commit()
    return ApiMessage(message="已向子任务执行人发送提醒")


@router.post("/tasks/{task_id}/members/{user_id}/task-remind", response_model=ApiMessage)
def remind_task_member(
    task_id: int,
    user_id: int,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiMessage:
    """按参与人发送任务提醒；有未完成子任务时聚焦子任务，否则按主任务提醒。"""

    task = db.query(Task).filter(Task.id == task_id, Task.deleted_at.is_(None)).first()
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    member = db.query(TaskMember).filter(TaskMember.task_id == task.id, TaskMember.user_id == user_id).first()
    if not member:
        raise HTTPException(status_code=404, detail="任务成员不存在")
    user = db.query(User).filter(User.id == user_id).first()
    subtasks = (
        db.query(TaskSubtask)
        .filter(TaskSubtask.task_id == task.id, TaskSubtask.assignee_id == user_id, TaskSubtask.status != "done")
        .order_by(TaskSubtask.sort_order.asc())
        .all()
    )
    if subtasks:
        assignee_name = user.name if user else f"成员#{user_id}"
        subtask_lines = []
        for index, item in enumerate(subtasks, start=1):
            line = f"{index}. {item.title}（执行人：{assignee_name}）"
            if item.content:
                line = f"{line}：{item.content}"
            subtask_lines.append(line)
        focus_text = f"请优先跟进 {len(subtasks)} 项未完成子任务"
        extra_context = {
            "remind_focus": focus_text,
            "subtask_summary": "\n".join(subtask_lines),
            "subtask_brief": "；".join(subtask_lines[:3]),
        }
    else:
        focus_text = f"请及时处理主任务“{task.title}”并反馈状态"
        extra_context = {"remind_focus": focus_text}
    _send_task_reminders(
        db,
        task,
        "manual_remind",
        recipient_user_ids=[user_id],
        extra_context=extra_context,
    )
    write_audit(
        db,
        current_user.id,
        "REMIND_TASK_MEMBER",
        "Task",
        task.id,
        {},
        {"user_id": user_id, "notification_created": True, "remind_focus": focus_text},
        module_name="api.task",
        message=f"向{user.name if user else user_id}发送任务提醒",
    )
    db.commit()
    return ApiMessage(message="已向该参与人发送提醒")


@router.post("/tasks/{task_id}/members/{user_id}/remind", response_model=ApiMessage)
def remind_task_member_channel(
    task_id: int,
    user_id: int,
    channel: str = Query(...),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiMessage:
    """按任务成员和指定渠道发送提醒，用于补发失败或未反馈的通知。"""

    safe_channel = channel.lower()
    if safe_channel not in {"email", "qax"}:
        raise HTTPException(status_code=400, detail="仅支持邮件或即时消息提醒")
    task = db.query(Task).filter(Task.id == task_id, Task.deleted_at.is_(None)).first()
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    member = db.query(TaskMember).filter(TaskMember.task_id == task.id, TaskMember.user_id == user_id).first()
    if not member:
        raise HTTPException(status_code=404, detail="任务成员不存在")
    user = db.query(User).filter(User.id == user_id).first()
    focus_text = f"请及时处理任务“{task.title}”并反馈状态"
    _send_task_reminders(
        db,
        task,
        "manual_remind",
        recipient_user_ids=[user_id],
        channels=[safe_channel],
        extra_context={"remind_focus": focus_text},
    )
    write_audit(
        db,
        current_user.id,
        "REMIND_TASK_MEMBER_CHANNEL",
        "Task",
        task.id,
        {},
        {"user_id": user_id, "channel": safe_channel, "notification_created": True},
        module_name="api.task",
        message=f"向{user.name if user else user_id}发送{NOTIFICATION_CHANNEL_LABELS.get(safe_channel, safe_channel)}提醒",
    )
    db.commit()
    return ApiMessage(message=f"已发送{NOTIFICATION_CHANNEL_LABELS.get(safe_channel, safe_channel)}提醒")


@router.post("/tasks/{task_id}/milestones/{milestone_id}/remind", response_model=ApiMessage)
def remind_task_milestone(
    task_id: int,
    milestone_id: int,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiMessage:
    """按里程碑节点发送提醒，帮助成员聚焦当前阶段目标。"""
    task = db.query(Task).filter(Task.id == task_id, Task.deleted_at.is_(None)).first()
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    milestone = (
        db.query(TaskMilestone)
        .filter(TaskMilestone.id == milestone_id, TaskMilestone.task_id == task.id)
        .first()
    )
    if not milestone:
        raise HTTPException(status_code=404, detail="里程碑不存在")
    if milestone.status == "done":
        raise HTTPException(status_code=400, detail="已完成的里程碑无需再次提醒")

    planned_text = milestone.planned_at.strftime("%Y-%m-%d %H:%M") if milestone.planned_at else "未设置"
    focus_text = f"请围绕里程碑“{milestone.name}”推进，节点时间：{planned_text}"
    _send_task_reminders(
        db,
        task,
        "manual_remind",
        extra_context={"remind_focus": focus_text},
    )
    write_audit(
        db,
        current_user.id,
        "REMIND_MILESTONE",
        "TaskMilestone",
        milestone.id,
        {},
        {"task_id": task.id, "notification_created": True, "remind_focus": focus_text},
    )
    db.commit()
    return ApiMessage(message="已向任务成员发送里程碑提醒")


@router.post("/tasks/due-remind/run", response_model=ApiMessage)
def run_due_remind(_: User = Depends(require_admin), db: Session = Depends(get_db)) -> ApiMessage:
    count = create_due_reminders(db)
    db.commit()
    return ApiMessage(message=f"已创建 {count} 条到期提醒通知")


@router.post("/delay-requests", response_model=dict)
def create_delay_request(payload: DelayRequestCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> dict:
    """成员提交延期申请，并同步通知管理员审批。"""
    task = db.query(Task).filter(Task.id == payload.task_id, Task.deleted_at.is_(None)).first()
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    request_obj = DelayRequest(
        task_id=task.id,
        applicant_id=current_user.id,
        apply_reason=payload.apply_reason,
        original_deadline=task.end_at,
        proposed_deadline=payload.proposed_deadline,
    )
    db.add(request_obj)
    db.flush()
    admin_ids = [item.id for item in db.query(User).filter(User.role.in_(tuple(ADMIN_ROLES)), User.is_active.is_(True)).all()]
    extra_context = {
        "delay_request_id": request_obj.id,
        "applicant_name": current_user.name,
        "proposed_deadline": payload.proposed_deadline.strftime("%Y-%m-%d"),
        "apply_reason": payload.apply_reason,
    }
    # 延期审批属于管理员定向通知，因此这里显式指定收件人列表。
    create_notification_with_recipients(db, task.id, "email", "delay_approval", "", recipient_user_ids=admin_ids, extra_context=extra_context)
    create_notification_with_recipients(db, task.id, "qax", "delay_approval", "", recipient_user_ids=admin_ids, extra_context=extra_context)
    write_audit(db, current_user.id, "CREATE_DELAY_REQUEST", "DelayRequest", request_obj.id, {}, {"task_id": task.id})
    db.commit()
    return {"id": request_obj.id}


@router.get("/delay-requests/pending", response_model=list[dict])
def list_pending_delay_requests(_: User = Depends(require_admin), db: Session = Depends(get_db)) -> list[dict]:
    """返回延期审批工作台列表。"""
    result = []
    for item in db.query(DelayRequest).order_by(DelayRequest.id.desc()).all():
        task = db.query(Task).filter(Task.id == item.task_id).first()
        applicant = db.query(User).filter(User.id == item.applicant_id).first()
        result.append(
            {
                "id": item.id,
                "task_id": item.task_id,
                "task_title": task.title if task else "",
                "task_priority": task.priority if task else "",
                "task_start_at": task.start_at if task else None,
                "task_end_at": task.end_at if task else None,
                "applicant_id": item.applicant_id,
                "applicant_name": applicant.name if applicant else "",
                "applicant_email": applicant.email if applicant else "",
                "approval_status": item.approval_status,
                "approval_status_text": DELAY_STATUS_LABELS.get(item.approval_status, item.approval_status),
                "apply_reason": item.apply_reason,
                "original_deadline": item.original_deadline,
                "proposed_deadline": item.proposed_deadline,
                "approved_deadline": item.approved_deadline,
                "approve_remark": item.approve_remark,
                "decided_by_channel": item.decided_by_channel,
                "version": item.version,
                "created_at": item.created_at,
            }
        )
    return result


@router.post("/delay-requests/{delay_id}/approve", response_model=dict)
def decide_delay_request(delay_id: int, payload: DelayDecisionRequest, current_user: User = Depends(require_admin), db: Session = Depends(get_db)) -> dict:
    """管理员审批延期申请，并复用服务层保证幂等与版本控制。"""
    request_obj = db.query(DelayRequest).filter(DelayRequest.id == delay_id).first()
    if not request_obj:
        raise HTTPException(status_code=404, detail="延期申请不存在")
    result, updated = apply_delay_decision(
        db=db,
        request_obj=request_obj,
        admin_id=current_user.id,
        request_id=payload.request_id,
        action=payload.action,
        channel="web",
        version=payload.version,
        remark=payload.remark,
        approved_deadline=payload.approved_deadline,
    )
    db.commit()
    return {
        "status": result,
        "status_text": DELAY_STATUS_LABELS.get(updated.approval_status, updated.approval_status),
        "id": updated.id,
        "approval_status": updated.approval_status,
    }


@router.post("/tasks/import", response_model=dict)
async def import_tasks(
    file: UploadFile = File(...),
    confirm_duplicate: bool = Form(False),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> dict:
    """导入任务 Excel，并在高重叠时要求二次确认。"""
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持导入 .xlsx 格式文件")

    workbook_bytes = await file.read()
    if not workbook_bytes:
        raise HTTPException(status_code=400, detail="上传文件为空")

    try:
        workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Excel 文件解析失败：{exc}") from exc

    sheet = workbook["任务导入模板"] if "任务导入模板" in workbook.sheetnames else workbook.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="模板中没有可导入的数据行")

    use_legacy_template = len(rows[0]) >= 10
    if use_legacy_template:
        import_rows, row_signatures, row_samples = _collect_import_rows(rows)
    else:
        import_rows, row_signatures, row_samples = _collect_simple_import_rows(rows)
    overlap_count, overlap_rate, overlap_samples = _detect_import_overlap(db, row_signatures, row_samples)
    if overlap_count >= 3 and overlap_rate >= 0.6 and not confirm_duplicate:
        preview = TaskImportPreviewOut(
            message=f"检测到本次导入与历史导入有 {overlap_count} 条高度重叠，请确认是否继续导入。",
            needs_confirmation=True,
            overlap_count=overlap_count,
            overlap_rate=round(overlap_rate, 2),
            overlap_samples=overlap_samples,
        )
        return preview.dict()

    created_task_ids: list[int] = []
    failures: list[dict[str, object]] = []
    for index, row_data in import_rows:
        try:
            payload = (
                _build_task_create_from_import_row(row_data, db)
                if use_legacy_template
                else _build_task_create_from_simple_import_row(row_data, db)
            )
            task = _create_task_record(payload, current_user, db, source="import")
            task_id = task.id
            db.commit()
            created_task_ids.append(task_id)
        except Exception as exc:
            db.rollback()
            detail = exc.detail if isinstance(exc, HTTPException) else str(exc)
            failures.append(
                {
                    "row_number": index,
                    "title": str(row_data.get("title") or "").strip(),
                    "reason": detail,
                }
            )

    success_count = len(created_task_ids)
    failure_count = len(failures)
    _record_task_import_history(
        db=db,
        filename=file.filename or "未命名导入文件.xlsx",
        current_user=current_user,
        row_signatures=row_signatures,
        row_samples=row_samples,
        success_count=success_count,
        failure_count=failure_count,
        overlap_count=overlap_count,
        confirmed_duplicate=confirm_duplicate,
        failures=failures,
        overlap_samples=overlap_samples,
        workbook_bytes=workbook_bytes,
    )
    db.commit()
    for task_id in created_task_ids:
        _enqueue_task_created_notifications(task_id)
    return {
        "message": f"任务导入完成：成功 {success_count} 条，失败 {failure_count} 条",
        "needs_confirmation": False,
        "success_count": success_count,
        "failure_count": failure_count,
        "created_task_ids": created_task_ids,
        "overlap_count": overlap_count,
        "overlap_samples": overlap_samples,
        "failures": failures,
    }


@router.get("/reports/export")
def export_report(_: User = Depends(require_admin), db: Session = Depends(get_db)) -> StreamingResponse:
    csv_buffer = io.StringIO()
    writer = csv.writer(csv_buffer)
    writer.writerow(["task_id", "title", "status", "completed_at", "delay_days", "subtask_count", "subtask_summary"])
    for task in db.query(Task).filter(Task.deleted_at.is_(None)).order_by(Task.id.asc()).all():
        subtasks = sorted(task.subtasks, key=lambda item: item.sort_order)
        subtask_summary = "；".join(f"{item.title}({item.assignee.name if item.assignee else item.assignee_id})" for item in subtasks)
        writer.writerow([task.id, task.title, task.main_status, task.completed_at or "", task.delay_days, len(subtasks), subtask_summary])
    return StreamingResponse(iter([csv_buffer.getvalue().encode("utf-8")]), media_type="text/csv")


@router.get("/audit-logs", response_model=list[AuditOut])
def list_audit_logs(_: User = Depends(require_admin), db: Session = Depends(get_db)) -> list[AuditOut]:
    user_name_map = {item.id: item.name for item in db.query(User).all()}
    return [
        AuditOut(
            id=item.id,
            log_level=getattr(item, "log_level", "INFO"),
            module_name=getattr(item, "module_name", "system"),
            action_type=item.action_type,
            message=getattr(item, "message", item.action_type),
            operator_id=item.operator_id,
            operator_name=user_name_map.get(item.operator_id or 0, ""),
            target_type=item.target_type,
            target_id=item.target_id,
            source_ip=item.source_ip,
            before_json=getattr(item, "before_json", "{}") or "{}",
            after_json=getattr(item, "after_json", "{}") or "{}",
            detail_json=getattr(item, "detail_json", "{}") or "{}",
            created_at=item.created_at,
        )
        for item in db.query(AuditLog).order_by(AuditLog.id.desc()).all()
    ]


@router.get("/system-logs", response_model=list[AuditOut])
def list_system_logs(_: User = Depends(require_admin), db: Session = Depends(get_db)) -> list[AuditOut]:
    """系统日志新入口，兼容旧的 `/audit-logs` 路由。"""
    return list_audit_logs(_, db)


@router.post("/system-logs/cleanup", response_model=ApiMessage)
def cleanup_expired_system_logs(current_user: User = Depends(require_admin), db: Session = Depends(get_db)) -> ApiMessage:
    """手动清理超过保留期的系统日志。"""
    deleted_count = cleanup_system_logs(db, settings.system_log_retention_days)
    write_audit(
        db,
        current_user.id,
        "CLEANUP_SYSTEM_LOG",
        "SystemLog",
        None,
        {},
        {"deleted_count": deleted_count, "retention_days": settings.system_log_retention_days},
        module_name="api.system-log",
        message=f"手动清理系统日志，删除 {deleted_count} 条过期记录",
        detail={"retention_days": settings.system_log_retention_days},
    )
    db.commit()
    return ApiMessage(message=f"已清理 {deleted_count} 条超过 {settings.system_log_retention_days} 天的系统日志")
